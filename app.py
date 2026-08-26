"""
JLT Inside-Sales Configurator - rough first-draft prototype.

Screens, built against the single JLT VMT price list ingested in
ingest/parse_vmt.py:

  /technical  - technical reviewer checkbox-approves which options are valid
  /sales      - rep picks a platform + approved options, saves/locks/prints/
                copies quotes tied to a manually-entered Opportunity ID
  /purchasing - all 4 price tiers, editable blanks, and an action-item report
                of what sales has quoted that still needs purchasing's input

Data lives in plain JSON files under data/ - no database yet, matching the
project's "start spreadsheet-based" approach. Cost and Current Cost are
purchasing-internal only and are never included in anything a rep or
customer sees (Sales screen, print view, Excel export).

Quote ID/lock rules:
  - A quote gets its Opportunity-Quote# the first time it's saved. Rev starts at 0.
  - Every successful save locks the quote (added 2026-08-17, per the user -
    was previously a separate manual toggle/Print-only). Locked "fixes" the
    quote - it can't be edited again until unlocked (Unlock is immediately
    clickable right after a save, no extra step needed).
  - Saving an already-saved quote bumps Rev by 1 whenever the configuration
    (selections/brand/platform) actually changed from what's stored -
    independent of lock state (every save locks now, so "only if ever
    locked" became meaningless the moment auto-lock shipped; replaced with
    a real content-diff check). A no-op re-save (nothing actually
    different) does not bump Rev.
  - Copy clones the configuration onto a new Opportunity ID, with a fresh
    Quote# and Rev back at 0.
"""

import copy
import csv
import io
import json
import re
import secrets
import sys
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

from flask import Flask, jsonify, redirect, render_template, request, send_file, session, url_for
import openpyxl
from openpyxl.styles import Font

import hubspot_client

BASE_DIR = Path(__file__).parent
sys.path.insert(0, str(BASE_DIR / "ingest"))
import parse_vmt        # noqa: E402
import parse_winmate     # noqa: E402
import parse_getac       # noqa: E402
import parse_cipherlab   # noqa: E402
from storage_facets import extract_storage_capacity, extract_storage_technology  # noqa: E402
from os_facets import extract_os_version, extract_os_edition  # noqa: E402
from cpu_facets import normalize_cpu_label, cpu_sort_key  # noqa: E402
from wwan_facets import (  # noqa: E402
    extract_wwan_generation, extract_wwan_module, extract_wwan_carrier, wwan_generation_sort_key,
)
from wifi_facets import extract_wifi_radio  # noqa: E402

# Brand -> parser, so Technical's upload form can route each vendor's
# spreadsheet to the parser that actually understands its layout instead of
# always assuming JLT's. Every entry here is a manufacturer's own official
# catalog (see is_selectable()) - third-party add-on vendors (RAM Mounts,
# Gamber-Johnson, etc.) aren't in this registry yet; that's a separate,
# not-yet-built ingestion path that will need requires_review=True instead.
PARSERS = {
    "JLT": parse_vmt.parse_workbook,
    "Winmate": parse_winmate.parse_workbook,
    "Getac": parse_getac.parse_workbook,
    "CipherLab": parse_cipherlab.parse_workbook,
}

app = Flask(__name__)

# Purchasing's "Save Prices" form has one row per flagged part (currently
# 3,421) x 5 fields each (row_key + 4 price fields) = ~17,100 form fields -
# real, legitimate data from a trusted internal user, not an attack. Found
# 2026-08-18 when the user hit a live 413 clicking Generate Catalog
# Report: Werkzeug 3.x's default safety limits (max_form_parts=1000,
# max_form_memory_size=500_000 bytes) reject the request before app.py
# ever sees it - Save Prices itself was equally broken for the full table,
# not just the report button. Raised well above today's catalog size, with
# headroom for it to grow.
app.config["MAX_FORM_PARTS"] = 50_000
app.config["MAX_FORM_MEMORY_SIZE"] = 5_000_000


def _get_app_version():
    """Short git commit hash the running app was started from - shown in the
    nav bar so it's obvious at a glance whether this instance matches what's
    on GitHub, given main can move independently of any given running
    process (see HANDOFF.md's review-checkpoint section). Falls back to
    "dev" if git isn't available (e.g. a zip export with no .git)."""
    import subprocess
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=BASE_DIR, capture_output=True, text=True, timeout=5, check=True,
        )
        return result.stdout.strip()
    except Exception:
        return "dev"


APP_VERSION = _get_app_version()


@app.context_processor
def inject_app_version():
    return {"app_version": APP_VERSION}


DATA_DIR = BASE_DIR / "data"
PARTS_FILE = DATA_DIR / "parts_vmt_q1_2026.json"
APPROVALS_FILE = DATA_DIR / "approvals.json"
QUOTES_FILE = DATA_DIR / "quotes.json"
CUSTOMERS_FILE = DATA_DIR / "customers.json"
SALES_REPS_FILE = DATA_DIR / "sales_reps.json"
PURCHASING_WARNINGS_FILE = DATA_DIR / "purchasing_warnings.json"
SITE_ACCESS_FILE = DATA_DIR / "site_access.json"
REPORTS_DIR = DATA_DIR / "reports"
# Holds the already-parsed rows of an uploaded pricing sheet between the
# preview step (dry-run counts shown, nothing saved) and Continue (real
# merge + save) - see purchasing()'s upload handling. Not cleaned up on a
# timer; an abandoned preview just leaves a small orphaned JSON file, same
# as REPORTS_DIR already does for generated reports nobody downloaded.
PENDING_IMPORTS_DIR = DATA_DIR / "pending_imports"

# --------------------------------------------------------------- site access
# A single shared PIN gating every page/API on the whole app - not per-user,
# just a soft "don't let a random person who finds the tunnel link poke
# around" gate. Same "not real security" caveat as the sales-rep codes: a
# short PIN with no lockout. Lives in a gitignored file (never committed -
# this repo is PUBLIC on GitHub) that's auto-created with a fresh random PIN
# and session secret key the first time the app runs.

def load_or_create_site_access():
    if SITE_ACCESS_FILE.exists():
        data = json.loads(SITE_ACCESS_FILE.read_text(encoding="utf-8"))
        # Backfill for files created before the Purchasing PIN existed
        # (2026-08-18) - an existing install's site_access.json predates
        # this key, so a plain load would KeyError on ["purchasing_pin"]
        # elsewhere without this.
        if "purchasing_pin" not in data:
            data["purchasing_pin"] = "1111"
            SITE_ACCESS_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")
        return data
    data = {
        "pin": "".join(secrets.choice("0123456789") for _ in range(4)),
        "secret_key": secrets.token_hex(32),
        # Fixed default (not random like the site PIN) per the user
        # (2026-08-18) - changeable from Admin same as the site PIN.
        "purchasing_pin": "1111",
    }
    SITE_ACCESS_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")
    return data


def save_site_access(data):
    SITE_ACCESS_FILE.write_text(json.dumps(data, indent=2), encoding="utf-8")


_site_access = load_or_create_site_access()
app.secret_key = _site_access["secret_key"]
# Governs both the site-wide session and the Purchasing session (same cookie,
# one lifetime - session.permanent is set once at site login, purchasing
# access just adds a flag to that same session rather than starting a
# separate one). Flask refreshes this on every request by default, so in
# practice it's "expires N days after the last visit," not N days after
# login. Reduced from 7 to 5 days per the user (2026-08-18).
app.permanent_session_lifetime = timedelta(days=5)


@app.before_request
def require_site_pin():
    if request.endpoint in ("login", "static"):
        return
    if not session.get("authenticated"):
        return redirect(url_for("login", next=request.path))


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    next_url = request.values.get("next") or url_for("sales")
    if request.method == "POST":
        entered = request.form.get("pin", "").strip()
        current = load_or_create_site_access()["pin"]
        if entered and entered == current:
            session.permanent = True
            session["authenticated"] = True
            return redirect(request.form.get("next") or url_for("sales"))
        error = "Incorrect PIN."
    return render_template("login.html", error=error, next=next_url)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# ---------------------------------------------------------- purchasing access
# A second, inner PIN gate scoped to just the Purchasing section - per the
# user (2026-08-18), even someone who already has the site-wide PIN
# shouldn't see Purchasing (cost data, purchasing-internal pricing) without
# this second code too. Same "not real security" caveat as the site PIN and
# sales-rep codes. Runs as its own before_request AFTER require_site_pin
# (registered below it) - Flask runs before_request handlers in registration
# order and stops at the first one that returns a response, so the site PIN
# is always checked first; this one only ever runs for an already
# site-authenticated session.
@app.before_request
def require_purchasing_pin():
    if request.endpoint in ("purchasing_login", "static", "login"):
        return
    if request.path.startswith("/purchasing") and not session.get("purchasing_authenticated"):
        return redirect(url_for("purchasing_login", next=request.path))


@app.route("/purchasing/login", methods=["GET", "POST"])
def purchasing_login():
    error = None
    next_url = request.values.get("next") or url_for("purchasing")
    if request.method == "POST":
        entered = request.form.get("pin", "").strip()
        current = load_or_create_site_access()["purchasing_pin"]
        if entered and entered == current:
            session["purchasing_authenticated"] = True
            return redirect(request.form.get("next") or url_for("purchasing"))
        error = "Incorrect PIN."
    return render_template("purchasing_login.html", error=error, next=next_url)


CATEGORY_ORDER = [
    "Base Unit:",
    "Processor Options",
    "RAM Memory Options:",
    "Storage Drive Options:",
    # "Storage Capacity"/"Storage Technology"/"OS Version"/"OS Edition" are
    # search-only pseudo-categories (see FACET_CATEGORIES below) - never a
    # real part's `category` value, only used to order/render them in the
    # Search by Requirements modal via category_sort_key(). Real Sales-page
    # category grouping/sorting still uses "Storage Drive Options:" and
    # "Operating System:" below, unaffected.
    "Storage Capacity",
    "Storage Technology",
    "Display options:",
    "Internal Options:",
    "Add On Options:",
    "IP Rating Options:",
    "Power Cable Options:",
    "Internal Wireless",
    "WWAN Generation",
    "WWAN Card",
    "WWAN Carrier",
    "Operating System:",
    "OS Version",
    "OS Edition",
]

PRICE_FIELDS = ["Floor Price", "MSRP", "Cost", "Current Cost"]

# Fixed-SKU brands (Getac, CipherLab - see ingest/parse_getac.py and
# ingest/parse_cipherlab.py) don't decompose into real per-category options,
# so there's nothing for Search by Requirements to match against for them
# under normal category/description matching. Their CPU/OS/RAM/etc info
# lives on the Base Unit row's `attributes` dict instead - this maps a
# searchable category to the attribute key that can satisfy it as a
# fallback. "Storage Capacity"/"Storage Technology" and "OS Version"/
# "OS Edition" replace what used to be single "Storage Drive Options:" and
# "Operating System:" entries (see FACET_CATEGORIES below and
# ingest/storage_facets.py's + ingest/os_facets.py's docstrings for why).
ATTRIBUTE_CATEGORY_MAP = {
    "Processor Options": "cpu",
    "RAM Memory Options:": "ram",
    "Storage Capacity": "storage",
    "Storage Technology": "storage_tech",
    "OS Version": "os_version",
    "OS Edition": "os_edition",
    "Display options:": "display",
    "Internal Wireless": "wireless",
}

# JLT/Winmate's real "Storage Drive Options:"/"Operating System:"/
# "Processor Options"/"Internal Wireless" rows have no precomputed
# attributes (unlike Getac) - a search on one of these synthetic categories
# has to derive the facet from each option's free-text description on the
# fly. Maps synthetic search category -> (list of real categories its
# options actually live under, list of extractor functions) - see
# ingest/storage_facets.py, ingest/os_facets.py, ingest/cpu_facets.py,
# ingest/wifi_facets.py, ingest/wwan_facets.py. Storage/OS split one real
# category into two synthetic ones; "Processor Options" maps to itself -
# it's not a split, just deduping near-identical spellings of the same real
# chip.
#
# The four wireless synthetic categories all share the same list of real
# source categories (_WIRELESS_REAL_CATEGORIES) rather than one real
# category each, because which real categories actually carry wireless data
# is brand-dependent: JLT's real "Internal Wireless" rows were split
# 2026-08-18 into three real categories - "Internal Wireless" (WiFi radio
# only), "WWAN Card", "WWAN Carrier" - one option per row exactly as
# before, just re-labeled by content (see the migration note in
# CHANGELOG.md). Winmate/Getac/CipherLab weren't touched and still have
# everything under real "Internal Wireless" alone. Scanning all three real
# categories for every synthetic one means e.g. a JLT row that's now real
# category "WWAN Card" (because it names a cellular module) but *also*
# mentions "AX210" in its description still shows up under the "Internal
# Wireless" WiFi facet too - nothing found via one field became unfindable
# via another just because of which real category a row's price/code now
# lives under.
_WIRELESS_REAL_CATEGORIES = ["Internal Wireless", "WWAN Card", "WWAN Carrier"]

FACET_CATEGORIES = {
    "Storage Capacity": (["Storage Drive Options:"], [extract_storage_capacity]),
    "Storage Technology": (["Storage Drive Options:"], [extract_storage_technology]),
    "OS Version": (["Operating System:"], [extract_os_version]),
    "OS Edition": (["Operating System:"], [extract_os_edition]),
    "Processor Options": (["Processor Options"], [normalize_cpu_label]),
    "Internal Wireless": (_WIRELESS_REAL_CATEGORIES, [extract_wifi_radio]),
    "WWAN Generation": (_WIRELESS_REAL_CATEGORIES, [extract_wwan_generation]),
    "WWAN Card": (_WIRELESS_REAL_CATEGORIES, [extract_wwan_module]),
    "WWAN Carrier": (_WIRELESS_REAL_CATEGORIES, [extract_wwan_carrier]),
}

# Reverse index: which real categories need facet-splitting instead of
# being pooled as an exact-match description, and into which synthetic
# categories each one splits.
_REAL_CATEGORY_TO_FACETS = {}
for _synthetic_cat, (_real_cats, _extractors) in FACET_CATEGORIES.items():
    for _real_cat in _real_cats:
        _REAL_CATEGORY_TO_FACETS.setdefault(_real_cat, []).append((_synthetic_cat, _extractors))


def _capacity_sort_key(label):
    """Numeric sort for capacity labels ("16GB", "1TB", ...) - plain string
    sort would put "1TB" between "128GB" and "16GB"."""
    m = re.match(r"(\d+)(GB|TB)", label)
    if not m:
        return (1, 0)
    n, unit = int(m.group(1)), m.group(2)
    return (0, n * 1024 if unit == "TB" else n)


def _os_version_sort_key(label):
    """Groups by OS family, then sorts numerically ascending within each
    family ("Android 9" before "Android 11" - plain string sort would put
    "Android 11" before "Android 9" since '1' < '9' character-by-character).
    """
    parts = label.split()
    family = parts[0] if parts else label
    version = 0.0
    for tok in parts[1:]:
        try:
            version = float(tok)
            break
        except ValueError:
            continue
    return (family, version)


CUSTOMER_FACING_PRICE_FIELDS = ["Floor Price", "MSRP"]  # Cost/Current Cost never leave Purchasing

# CipherLab was excluded here 2026-08-17 because its old source spreadsheet
# was a price-*increase* list, not a full catalog - ~160 search-dropdown
# values had no Base Unit row to ever match. Resolved 2026-08-25: the old
# CipherLab source/data was replaced entirely with a new price-book workbook
# (see ingest/parse_cipherlab.py) where every ingested platform has real
# Base Unit rows with real attributes, and api_search_options already
# excludes any (brand, platform) with zero Base Unit rows from the dropdown
# pool - so nothing here needs a brand-wide exclusion anymore. Kept as an
# empty set (rather than deleted outright) so a future vendor with the same
# "spreadsheet isn't a full catalog" problem has an established mechanism to
# reuse instead of re-inventing it.
SEARCH_EXCLUDED_BRANDS = set()

# Fixed roster so the Brand dropdown always shows every vendor JLT resells,
# not just whichever ones happen to have data today. All 4 are ingested and
# auto-approved as of 2026-08-16 (see PARSERS above and each brand's
# requires_review=False) - this list existing independently of ingestion
# status just means a brand added here before its parser/data exist would
# show an empty state instead of being missing from the dropdown entirely.
BRANDS = ["JLT", "Winmate", "Getac", "CipherLab"]


def category_sort_key(category):
    normalized = category.strip().rstrip(":")
    for i, c in enumerate(CATEGORY_ORDER):
        if c.strip().rstrip(":") == normalized:
            return i
    return len(CATEGORY_ORDER)


def money_value(v):
    if v in (None, ""):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().upper()
    if s in ("INCL", "NC"):
        return 0.0
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return 0.0


# ---------------------------------------------------------------- parts data

def load_parts():
    return json.loads(PARTS_FILE.read_text(encoding="utf-8"))


def save_parts(parts):
    PARTS_FILE.write_text(json.dumps(parts, indent=2), encoding="utf-8")


def part_key(p):
    return (p["brand"], p["platform"], p["category"], p["code"])


def find_part(parts, brand, platform, category, code):
    for p in parts:
        if p["brand"] == brand and p["platform"] == platform and p["category"] == category and p["code"] == code:
            return p
    return None


def is_selectable(p, approvals):
    """A part is usable on Sales if it's from a manufacturer's own official
    catalog (requires_review=False - JLT/Winmate/Getac/CipherLab's own
    published price books today) or has been explicitly Technical-approved.
    Missing the field entirely defaults to requiring review (safe default -
    nothing slips through unreviewed by accident). Third-party add-ons (RAM
    Mounts, Gamber-Johnson, etc. - not built yet) will always need the
    explicit-approval path, since a mount vendor's own catalog doesn't
    self-certify compatibility with a specific host platform the way an
    OEM's own spec sheet does."""
    if not p.get("requires_review", True):
        return True
    return part_key(p) in approvals


# ------------------------------------------------------------------ approvals

def load_approvals():
    if not APPROVALS_FILE.exists():
        return set()
    raw = json.loads(APPROVALS_FILE.read_text(encoding="utf-8"))
    return {tuple(x) for x in raw}


def save_approvals(approvals):
    APPROVALS_FILE.write_text(
        json.dumps(sorted(list(t) for t in approvals), indent=2), encoding="utf-8"
    )


# ---------------------------------------------------------------- customers
# A local stand-in for a real CRM customer list until HubSpot is connected.
# Just names for now - enough to look up / create / attach to a quote.

def load_customers():
    if not CUSTOMERS_FILE.exists():
        return []
    return json.loads(CUSTOMERS_FILE.read_text(encoding="utf-8"))


def save_customers(customers):
    CUSTOMERS_FILE.write_text(json.dumps(customers, indent=2), encoding="utf-8")


# --------------------------------------------------------------- sales reps
# Lightweight rep identification for quote attribution - NOT real security.
# A 4-digit code (last 4 of cell number) is easily guessable; this exists so
# a quote records who touched it, not to gate access to anything.

def load_sales_reps():
    if not SALES_REPS_FILE.exists():
        return []
    return json.loads(SALES_REPS_FILE.read_text(encoding="utf-8"))


def save_sales_reps(reps):
    SALES_REPS_FILE.write_text(json.dumps(reps, indent=2), encoding="utf-8")


def rep_code_matches(rep_match, code):
    """A locked rep (Admin's "Lock" toggle) can't verify/save/copy with
    their code even if it's still correct - locking is meant to actually
    stop further use, not just hide the rep from the picker."""
    return rep_match is not None and not rep_match.get("locked") and rep_match["code"] == code


# ----------------------------------------------------------- purchasing warnings
# Standing notices for Purchasing that don't fit the per-part/per-quote
# flagging the Dashboard already does - things purchasing needs to be told
# once and acknowledge, not a per-row count. Added 2026-08-18, per the user,
# seeded with the finding that JLT's option codes don't automatically match
# Jeeves' USItem# (see CHANGELOG.md's Jeeves TODO entry the same day) - that
# kind of "here's a gap you need to know about" note is exactly what this is
# for. Auto-created on first run with that one seed warning, same pattern as
# load_or_create_site_access().

def load_or_create_purchasing_warnings():
    if PURCHASING_WARNINGS_FILE.exists():
        return json.loads(PURCHASING_WARNINGS_FILE.read_text(encoding="utf-8"))
    warnings = [{
        "id": secrets.token_hex(8),
        "message": (
            "Jeeves Part Number can't be auto-matched from the JLT price book. A real Jeeves "
            "export was checked against the catalog (2026-08-18): 0 of 85 JLT option codes "
            "matched Jeeves' USItem# exactly, and only 3 of 125 descriptions matched - Jeeves "
            "tracks much more granular internal part numbers (e.g. CB-00639-50) than JLT's "
            "short price-book codes (e.g. \"H\", \"KA\"). Jeeves Part Numbers will need to be "
            "assigned by hand, part by part - there's no bulk import that can do this safely."
        ),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "acknowledged": False,
        "acknowledged_at": None,
    }]
    save_purchasing_warnings(warnings)
    return warnings


def save_purchasing_warnings(warnings):
    PURCHASING_WARNINGS_FILE.write_text(json.dumps(warnings, indent=2), encoding="utf-8")


# --------------------------------------------------------------------- quotes

def load_quotes():
    if not QUOTES_FILE.exists():
        return {}
    return json.loads(QUOTES_FILE.read_text(encoding="utf-8"))


def save_quotes(quotes):
    QUOTES_FILE.write_text(json.dumps(quotes, indent=2), encoding="utf-8")


def lineage_key(opportunity_id, quote_number):
    return f"{opportunity_id}::{quote_number}"


def display_id(q):
    return f"{q['opportunity_id']}-{q['quote_number']}-{q['rev_number']}"


def revision_snapshot(q):
    """Captures everything about a quote that actually changes between
    revisions - added 2026-08-25 to fix the known limitation that a new
    revision used to overwrite the previous one's data in place, with
    nothing kept to browse back through (see CHANGELOG.md's Pending/TODO).
    Deliberately excludes `locked` - that's a current-quote-level concept,
    not something meaningful to track per historical revision."""
    return {
        "rev_number": q["rev_number"],
        "customer": q.get("customer", ""),
        "brand": q.get("brand", "JLT"),
        "platform": q["platform"],
        "selections": q["selections"],
        "floor_total": q["floor_total"],
        "msrp_total": q["msrp_total"],
        "part_number": q["part_number"],
        "sales_rep": q.get("sales_rep", ""),
        "updated_at": q["updated_at"],
    }


def next_quote_number(quotes, opportunity_id):
    nums = [q["quote_number"] for q in quotes.values() if q["opportunity_id"] == opportunity_id]
    return max(nums, default=0) + 1


def build_snapshot(parts, brand, platform, selections):
    """selections: {category: code} -> list of line items with Floor/MSRP only
    (never Cost). A category's value can also be a list of codes (added
    2026-08-26, for MULTI_SELECT_CATEGORIES on the client - e.g. Add On
    Options:, where a rep may legitimately want more than one) - each code
    in the list becomes its own line, same shape as a single-code category."""
    lines = []
    for category, codes in selections.items():
        code_list = codes if isinstance(codes, list) else [codes]
        for code in code_list:
            part = find_part(parts, brand, platform, category, code)
            if part is None:
                continue
            lines.append({
                "category": category,
                "code": code,
                "description": part.get("description"),
                "Floor Price": part.get("Floor Price"),
                "MSRP": part.get("MSRP"),
            })
    lines.sort(key=lambda l: category_sort_key(l["category"]))
    return lines


def quote_totals(lines):
    floor_total = sum(money_value(l["Floor Price"]) for l in lines)
    msrp_total = sum(money_value(l["MSRP"]) for l in lines)
    return floor_total, msrp_total


def quote_part_number(platform, lines):
    return "-".join([platform] + [l["code"] for l in lines])


def compute_quote_action_items(parts, quotes):
    """Line items in saved quotes whose part-level data is still incomplete in
    the *current* parts data (not the quote's frozen snapshot): a missing
    Jeeves Part Number, Floor Price, MSRP, Cost, or Current Cost. Originally
    just Cost/Current Cost - extended 2026-08-18, per the user, once
    Technical gained the ability to add a not-yet-fully-priced option
    directly (see "Add Option" on Technical): a technician-added option can
    get quoted before Purchasing has assigned it a real Jeeves Part Number
    or any price at all, and this is exactly what's meant to catch that.
    Shared by the Purchasing report, the Purchasing dashboard counters, and
    the Admin counter so none of them ever drift apart from each other."""
    action_items = []
    for q in quotes.values():
        brand = q.get("brand", "JLT")
        for line in q["selections"]:
            part = find_part(parts, brand, q["platform"], line["category"], line["code"])
            missing = []
            if not part or not part.get("jeeves_part_number"):
                missing.append("Jeeves Part #")
            for f in ("Floor Price", "MSRP", "Cost", "Current Cost"):
                if not part or part.get(f) in (None, ""):
                    missing.append(f)
            if missing:
                action_items.append({
                    "display_id": display_id(q),
                    "opportunity_id": q["opportunity_id"],
                    "locked": q["locked"],
                    "brand": brand,
                    "platform": q["platform"],
                    "category": line["category"],
                    "code": line["code"],
                    "description": line["description"],
                    "missing": ", ".join(missing),
                    "missing_fields": missing,
                })
    return action_items


def compute_quote_action_item_counts(action_items):
    """Per-field breakdown of compute_quote_action_items() for the Purchasing
    dashboard - "N quoted line items missing X" for each field independently,
    since a single line item can be missing more than one field at once."""
    counts = {"Jeeves Part #": 0, "Floor Price": 0, "MSRP": 0, "Cost": 0, "Current Cost": 0}
    for item in action_items:
        for f in item["missing_fields"]:
            counts[f] += 1
    return counts


def compute_unreviewed_base_models(parts, approvals):
    """Platforms whose Base Unit option (the platform's base config) hasn't
    been technical-approved yet - i.e. review on that platform hasn't even
    started. Auto-approved (manufacturer-catalog) base units never need
    review in the first place, so they're excluded rather than counted as
    outstanding work."""
    unreviewed = []
    for p in parts:
        if p["category"] != "Base Unit:":
            continue
        if not p.get("requires_review", True):
            continue
        if part_key(p) not in approvals:
            unreviewed.append(p)
    unreviewed.sort(key=lambda p: (p["brand"], p["platform"]))
    return unreviewed


def compute_unlinked_customers(customers):
    """Manually-created customers with no hubspot_id yet - an action list so
    a manually-typed name doesn't quietly stay disconnected from the real
    HubSpot record once that connector exists."""
    unlinked = [c for c in customers if c.get("source") == "manual" and not c.get("hubspot_id")]
    unlinked.sort(key=lambda c: c["created_at"])
    return unlinked


# ------------------------------------------------------------- spreadsheet upload

def merge_parts(parts, new_rows, allow_new):
    """Merge parsed spreadsheet rows into the working parts list, matched by
    (brand, platform, category, code). A blank cell in the upload never
    erases a value already on file - only a non-blank incoming value
    overwrites, so a partial vendor refresh can't wipe out prices purchasing
    already filled in. allow_new=False means unmatched rows are skipped
    rather than creating a new part - used for Purchasing's pricing-only
    upload, which shouldn't be able to invent new catalog entries (that's
    Technical's job). Returns (added, updated, skipped) counts."""
    added = updated = skipped = 0
    # "requires_review" and "attributes" merge the same way as description/price:
    # a blank/missing incoming value never erases what's already on file. Every
    # brand parser sets requires_review explicitly (True/False, never blank), so
    # re-uploading an OEM catalog through Technical carries the same
    # auto-approval policy as the initial ingest, rather than silently
    # defaulting re-uploaded rows back to "needs review".
    mergeable_fields = ["description", "requires_review", "attributes"] + PRICE_FIELDS
    BLANK = (None, "", {}, [])

    for row in new_rows:
        brand = (row.get("brand") or "JLT").strip()
        platform = (row.get("platform") or "").strip()
        category = (row.get("category") or "").strip()
        code = (row.get("code") or "").strip() if row.get("code") is not None else ""
        if not (platform and category and code):
            skipped += 1
            continue

        existing = find_part(parts, brand, platform, category, code)
        if existing is None:
            if not allow_new:
                skipped += 1
                continue
            new_part = {"brand": brand, "platform": platform, "category": category, "code": code, "description": None, "requires_review": True}
            for f in PRICE_FIELDS:
                new_part[f] = None
            for f in mergeable_fields:
                if row.get(f) not in BLANK:
                    new_part[f] = row[f]
            parts.append(new_part)
            added += 1
        else:
            changed = False
            for f in mergeable_fields:
                val = row.get(f)
                if val not in BLANK and existing.get(f) != val:
                    existing[f] = val
                    changed = True
            if changed:
                updated += 1

    return added, updated, skipped


def parse_flat_price_table(file_storage):
    """Read Purchasing's flat pricing format - the same columns as its own
    'Generate Catalog Report': brand, platform, category, code, description,
    Floor Price, MSRP, Cost, Current Cost - from an uploaded .xlsx or .csv.
    A missing/blank brand column defaults to JLT (old reports predate the
    brand column)."""
    filename = (file_storage.filename or "").lower()
    rows = []

    if filename.endswith(".csv"):
        text = file_storage.read().decode("utf-8-sig")
        for r in csv.DictReader(io.StringIO(text)):
            rows.append(r)
    else:
        wb = openpyxl.load_workbook(file_storage, data_only=True)
        ws = wb.active
        header_cells = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [c.value for c in header_cells]
        for values in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(headers, values)))

    out = []
    for r in rows:
        out.append({
            "brand": r.get("brand") or "JLT",
            "platform": r.get("platform"),
            "category": r.get("category"),
            "code": r.get("code"),
            "description": r.get("description"),
            "Floor Price": r.get("Floor Price"),
            "MSRP": r.get("MSRP"),
            "Cost": r.get("Cost"),
            "Current Cost": r.get("Current Cost"),
        })
    return out


# --------------------------------------------------------------------- pages

@app.route("/")
def index():
    return redirect(url_for("sales"))


@app.route("/technical", methods=["GET", "POST"])
def technical():
    upload_result = None
    add_option_error = None
    add_option_result = None

    # Which single brand's platforms/options are shown - the catalog is
    # 3,551 parts across 4 brands now, too much to usefully show at once (and
    # confusing: picking JLT should mean *only* JLT, not JLT mixed in with
    # every other brand on the same scroll). Carried as "view" (not "brand",
    # which the upload form below already uses for a different purpose - the
    # upload target brand, independent of which brand is currently shown).
    selected_brand = request.values.get("view") or "JLT"
    if selected_brand not in BRANDS:
        selected_brand = "JLT"

    if request.method == "POST":
        uploaded = request.files.get("file")
        if uploaded and uploaded.filename:
            brand = request.form.get("brand", "JLT")
            parser = PARSERS.get(brand)
            if parser is None:
                upload_result = {"error": f"No parser registered for brand {brand!r} yet."}
            else:
                try:
                    new_rows = parser(uploaded, brand=brand)
                except Exception as e:
                    upload_result = {"error": str(e)}
                else:
                    parts = load_parts()
                    added, updated, skipped = merge_parts(parts, new_rows, allow_new=True)
                    save_parts(parts)
                    upload_result = {"added": added, "updated": updated, "skipped": skipped, "total": len(new_rows), "brand": brand}
        elif request.form.get("action") == "add_option":
            # Lets a technician add one new valid option JLT engineering has
            # signed off on but that isn't in the current vendor price book
            # snapshot - e.g. an add-on accessory qualified after the last
            # spreadsheet ingest. Always requires_review=True (same flag
            # third-party add-ons will use once that path exists) - it needs
            # the same checkbox approval as anything not from the official
            # catalog, and it's what lets Purchasing later notice it was
            # quoted before it had a real Jeeves Part Number/price (see
            # compute_quote_action_items()). Deliberately create-only - an
            # existing (brand, platform, category, code) is skipped, not
            # silently overwritten, unlike the bulk upload path above.
            # No price fields at all here, per the user (2026-08-18):
            # Technical never touches pricing - that's Purchasing's job,
            # and is exactly what the "unpriced, needs review" state and the
            # Purchasing Dashboard/warnings exist to surface.
            brand = request.form.get("add_brand", selected_brand)
            add_platforms = [p.strip() for p in request.form.getlist("add_platforms") if p.strip()]
            category = request.form.get("add_category", "").strip()
            code = request.form.get("add_code", "").strip()
            description = request.form.get("add_description", "").strip()
            jeeves_part_number = request.form.get("add_jeeves_part_number", "").strip() or None

            if not (add_platforms and category and code and description):
                add_option_error = "Vendor, at least one Platform, Category, Code, and Description are all required."
            else:
                parts = load_parts()
                added_platforms = []
                skipped_platforms = []
                for platform in add_platforms:
                    if find_part(parts, brand, platform, category, code) is not None:
                        skipped_platforms.append(platform)
                        continue
                    parts.append({
                        "brand": brand,
                        "platform": platform,
                        "category": category,
                        "code": code,
                        "description": description,
                        "Floor Price": None,
                        "MSRP": None,
                        "Cost": None,
                        "Current Cost": None,
                        "requires_review": True,
                        "jeeves_part_number": jeeves_part_number,
                    })
                    added_platforms.append(platform)
                if added_platforms:
                    save_parts(parts)
                add_option_result = {
                    "brand": brand, "category": category, "code": code, "description": description,
                    "added_platforms": added_platforms, "skipped_platforms": skipped_platforms,
                }
        elif "approved" in request.form or request.form.get("form") == "approvals":
            # Only this one brand's checkboxes were ever rendered (the page
            # is filtered to one brand - see selected_brand above), so a
            # naive "replace the whole approvals set with what was
            # submitted" would silently wipe out every other brand's
            # approvals. Replace only this brand's entries; leave every
            # other brand's approvals exactly as they were.
            approvals_brand = request.form.get("approvals_brand", selected_brand)
            selected = request.form.getlist("approved")
            new_approvals = {a for a in load_approvals() if a[0] != approvals_brand}
            for raw_key in selected:
                brand, platform, category, code = raw_key.split("||", 3)
                new_approvals.add((brand, platform, category, code))
            save_approvals(new_approvals)

    parts = load_parts()
    approvals = load_approvals()

    brands = {b: {} for b in BRANDS}
    for p in parts:
        platforms = brands.setdefault(p["brand"], {})
        platforms.setdefault(p["platform"], []).append(p)
    for platforms in brands.values():
        for plist in platforms.values():
            plist.sort(key=lambda p: (category_sort_key(p["category"]), p["code"] or ""))

    brands = {
        brand: dict(sorted(brands[brand].items()))
        for brand in BRANDS
    }

    # For "Add a new option": every brand's platform list (so the Platform
    # checkboxes can be swapped client-side when Vendor changes, without a
    # round trip) and the curated CATEGORY_ORDER vocabulary (not raw catalog
    # data - reverted 2026-08-19, per the user, after first trying the
    # opposite fix of adding every raw category to CATEGORY_ORDER, which
    # wrongly changed Search too). Some vendors' real ingested data contains
    # narrow, collision-prone pass-through categories (Winmate's CANBUS/
    # DIDO/LAN/Camera/Data Collection:/Data Collection: (2), JLT's Dock -
    # see ingest/category_map.py's docstring for why those were never
    # folded into the canonical vocabulary) that a technician should never
    # be offered as a category to add a *new* option under. Also excludes
    # CATEGORY_ORDER's own search-only pseudo-categories (Storage Capacity/
    # Storage Technology/WWAN Generation/OS Version/OS Edition - see the
    # comment above CATEGORY_ORDER) since no real part can ever have one of
    # those as its actual category.
    platforms_by_brand = {b: sorted(brands.get(b, {}).keys()) for b in BRANDS}
    _search_only_pseudo_categories = {
        "Storage Capacity", "Storage Technology", "WWAN Generation", "OS Version", "OS Edition",
    }
    all_categories = [c for c in CATEGORY_ORDER if c not in _search_only_pseudo_categories]

    return render_template(
        "technical.html",
        brands=brands,
        selected_brand=selected_brand,
        all_brands=BRANDS,
        approvals=approvals,
        upload_result=upload_result,
        platforms_by_brand=platforms_by_brand,
        all_categories=all_categories,
        add_option_error=add_option_error,
        add_option_result=add_option_result,
    )


@app.route("/sales")
def sales():
    parts = load_parts()
    approvals = load_approvals()
    approved_parts = [p for p in parts if is_selectable(p, approvals)]

    # {brand: {platform: {category: [options]}}} - Brand is the first choice
    # on the page, filtering which platforms show up, same as Platform then
    # filters which category dropdowns show up.
    brands = {}
    for p in approved_parts:
        platforms = brands.setdefault(p["brand"], {})
        platforms.setdefault(p["platform"], {}).setdefault(p["category"], []).append(p)

    for platforms in brands.values():
        for cats in platforms.values():
            for plist in cats.values():
                plist.sort(key=lambda p: p["code"] or "")

    brands_with_data = sorted(brands.keys())

    return render_template(
        "sales.html",
        all_brands=BRANDS,
        brands_with_data=brands_with_data,
        brands_json=json.dumps(brands),
        category_order=CATEGORY_ORDER,
        search_excluded_brands=SEARCH_EXCLUDED_BRANDS,
    )


@app.route("/api/search_options")
def api_search_options():
    """Distinct approved option descriptions per category, for populating the
    'Search by Requirements' dropdowns - a description (not a code) is the
    match key since the same code can mean different things on different
    platforms, but the description text is the actual spec a rep cares about.
    Optionally scoped to one brand via ?brand=."""
    brand_filter = request.args.get("brand") or None
    parts = load_parts()
    approvals = load_approvals()
    approved_parts = [p for p in parts if is_selectable(p, approvals) and p["brand"] not in SEARCH_EXCLUDED_BRANDS]

    # Some source spreadsheets aren't a full catalog - e.g. CipherLab's is a
    # *price increase* list, so a product family whose base unit price didn't
    # change can appear with only its accessories/warranties/software SKUs
    # and no "Base Unit:" row at all (confirmed 2026-08-17: 8600, HERA51, and
    # a batch of Wavelink/Ivanti license SKUs). An option under one of those
    # platforms can never resolve to a system - Search by Requirements finds
    # base units, not accessories - so it must not be offered as a
    # requirement in the first place, or picking it always returns zero
    # results no matter how correct the matching logic is.
    platforms_with_base_unit = {
        (p["brand"], p["platform"]) for p in approved_parts if p["category"] == "Base Unit:"
    }

    by_category = {}
    for p in approved_parts:
        if brand_filter and p["brand"] != brand_filter:
            continue
        if (p["brand"], p["platform"]) not in platforms_with_base_unit:
            continue

        if p["category"] == "Base Unit:":
            # Not a selectable option itself, but a fixed-SKU brand's only
            # source of CPU/OS/RAM info - surface those as if they were
            # options in the matching pseudo-category (see
            # ATTRIBUTE_CATEGORY_MAP) so Getac/CipherLab base units show up
            # in the same requirement search as JLT/Winmate's real options.
            for category, attr_key in ATTRIBUTE_CATEGORY_MAP.items():
                val = (p.get("attributes") or {}).get(attr_key)
                if val:
                    by_category.setdefault(category, set()).add(val)
            continue

        if p["category"] in _REAL_CATEGORY_TO_FACETS:
            # Split into independent, loosely-matched facets instead of one
            # flat exact-description dropdown - see FACET_CATEGORIES and
            # ingest/storage_facets.py's/ingest/os_facets.py's/
            # ingest/wifi_facets.py's/ingest/wwan_facets.py's docstrings for
            # why (per the user, 2026-08-17/18/19: a capacity search
            # shouldn't care about storage technology and vice versa; an OS
            # search shouldn't care about GAC/LTSC licensing channels; WWAN
            # generation/card/carrier should each be searchable
            # independently of each other and of the WiFi radio itself).
            # Every real category listed here is fully replaced by its
            # facet(s) - none of them fall through to the raw description
            # below.
            for category, extractors in _REAL_CATEGORY_TO_FACETS[p["category"]]:
                for extractor in extractors:
                    val = extractor(p.get("description"))
                    if val:
                        by_category.setdefault(category, set()).add(val)
            continue

        if not p.get("description"):
            continue
        by_category.setdefault(p["category"], set()).add(p["description"])

    out = {
        cat: sorted(descs, key=_capacity_sort_key) if cat == "Storage Capacity"
        else sorted(descs, key=_os_version_sort_key) if cat == "OS Version"
        else sorted(descs, key=cpu_sort_key) if cat == "Processor Options"
        else sorted(descs, key=wwan_generation_sort_key) if cat == "WWAN Generation"
        else sorted(descs)
        for cat, descs in by_category.items()
    }
    return jsonify(out)


@app.route("/api/search_base_units", methods=["POST"])
def api_search_base_units():
    """Finds every (brand, platform, Base Unit SKU) whose *approved* options
    satisfy every selected requirement (category -> exact description match).
    Requirements are optional and independent - a rep can fill in just
    Storage and OS and leave the rest blank, and any base unit with both
    stays in the results regardless of what else it does or doesn't offer.

    JLT/Winmate have exactly one Base Unit row per platform plus real,
    independently-selectable per-category option rows (Processor Options,
    RAM Memory Options, etc.) - a criterion on one of those categories is
    checked against every option row for the platform, since any of them
    can be combined freely with any Base Unit chassis choice.

    Getac/CipherLab have no such per-category rows - each "Base Unit:" row
    *is* a complete, already-fixed SKU, and a single platform (e.g. Getac's
    B360G3) can have several Base Unit rows, one per sellable SKU, each with
    its own cpu/os/ram/storage/display/wireless in `attributes`. A criterion
    on one of those pseudo-categories (see ATTRIBUTE_CATEGORY_MAP) must
    therefore be checked against *each* Base Unit row individually, and all
    criteria must be satisfied by the *same* row - checking against just one
    arbitrarily-picked row per platform (the previous approach) silently
    missed every SKU whose attributes weren't on that one row, and checking
    each criterion independently across different rows would wrongly match
    combinations no single real SKU actually offers."""
    body = request.get_json(force=True)
    brand_filter = body.get("brand") or None
    criteria = {k: v for k, v in (body.get("criteria") or {}).items() if v}

    parts = load_parts()
    approvals = load_approvals()
    approved_parts = [p for p in parts if is_selectable(p, approvals) and p["brand"] not in SEARCH_EXCLUDED_BRANDS]

    grouped = {}
    for p in approved_parts:
        grouped.setdefault((p["brand"], p["platform"]), []).append(p)

    def real_category_met(options, category, desc):
        if category in FACET_CATEGORIES:
            # JLT/Winmate's real "Storage Drive Options:"/"Operating
            # System:" rows have no precomputed facet attributes - derive
            # it from each option's description the same way
            # api_search_options does when building the dropdown, so the
            # two stay in sync.
            real_categories, extractors = FACET_CATEGORIES[category]
            return any(
                o["category"] in real_categories
                and any(extractor(o.get("description")) == desc for extractor in extractors)
                for o in options
            )
        return any(o["category"] == category and o.get("description") == desc for o in options)

    matches = []
    for (brand, platform), options in grouped.items():
        if brand_filter and brand != brand_filter:
            continue
        base_units = [o for o in options if o["category"] == "Base Unit:"]
        if not base_units:
            continue

        # Criteria satisfiable by a real, platform-wide option category
        # (JLT/Winmate) don't need to be re-checked per Base Unit row - any
        # Base Unit chassis choice can be paired with any of those options.
        remaining = {
            category: desc for category, desc in criteria.items()
            if not real_category_met(options, category, desc)
        }

        for bu in base_units:
            attrs = bu.get("attributes") or {}
            if all(
                attrs.get(ATTRIBUTE_CATEGORY_MAP.get(category)) == desc
                for category, desc in remaining.items()
            ):
                matches.append({
                    "brand": brand,
                    "platform": platform,
                    "code": bu.get("code"),
                    "description": bu.get("description"),
                    "floor_price": bu.get("Floor Price"),
                    "msrp": bu.get("MSRP"),
                })

    matches.sort(key=lambda m: (m["brand"], m["platform"], m["code"] or ""))
    return jsonify(matches)


@app.route("/quote/<path:opportunity_id>/<int:quote_number>/print")
def quote_print(opportunity_id, quote_number):
    quotes = load_quotes()
    key = lineage_key(opportunity_id, quote_number)
    q = quotes.get(key)
    if q is None:
        return "Quote not found", 404

    if not q["locked"]:
        q["locked"] = True
        q["ever_locked"] = True
        q["updated_at"] = datetime.now().isoformat(timespec="seconds")
        save_quotes(quotes)

    return render_template("quote_print.html", q=q, display_id=display_id(q))


@app.route("/purchasing", methods=["GET", "POST"])
def purchasing():
    parts = load_parts()
    quotes = load_quotes()
    action = request.form.get("action") if request.method == "POST" else None

    # Upload now previews instead of applying immediately (added 2026-08-18,
    # per the user - the old immediate-apply behavior gave no chance to
    # catch a bad file before it was already saved). Parses the file and
    # dry-runs merge_parts() against a deep copy to get real would-be
    # counts without touching the live parts list, then stashes the
    # already-parsed rows as JSON so Continue doesn't need to re-parse the
    # original CSV/XLSX - see PENDING_IMPORTS_DIR.
    upload_result = None
    pending_import = None
    uploaded = request.files.get("file") if request.method == "POST" else None
    if uploaded and uploaded.filename:
        try:
            new_rows = parse_flat_price_table(uploaded)
        except Exception as e:
            upload_result = {"error": str(e)}
        else:
            _added, updated, skipped = merge_parts(copy.deepcopy(parts), new_rows, allow_new=False)
            PENDING_IMPORTS_DIR.mkdir(parents=True, exist_ok=True)
            token = secrets.token_hex(8)
            (PENDING_IMPORTS_DIR / f"{token}.json").write_text(json.dumps(new_rows), encoding="utf-8")
            pending_import = {"token": token, "updated": updated, "skipped": skipped, "total": len(new_rows)}

    if action == "confirm_import":
        token = request.form.get("token", "")
        pending_path = PENDING_IMPORTS_DIR / f"{token}.json"
        if pending_path.is_file():
            new_rows = json.loads(pending_path.read_text(encoding="utf-8"))
            added, updated, skipped = merge_parts(parts, new_rows, allow_new=False)
            save_parts(parts)
            parts = load_parts()
            pending_path.unlink()
            upload_result = {"added": added, "updated": updated, "skipped": skipped, "total": len(new_rows)}
        else:
            upload_result = {"error": "That import preview has expired or was already applied."}

    if action == "cancel_import":
        token = request.form.get("token", "")
        pending_path = PENDING_IMPORTS_DIR / f"{token}.json"
        if pending_path.is_file():
            pending_path.unlink()
        upload_result = {"cancelled": True}

    jeeves_message = None
    if action == "jeeves_part_compare":
        jeeves_message = (
            "Jeeves isn't connected yet. This button is a placeholder for that future "
            "integration - once available, it will check every catalog part's assigned "
            "Jeeves Part Number against Jeeves and list any part not recognized there."
        )
    if action == "jeeves_price_compare":
        jeeves_message = (
            "Jeeves isn't connected yet. This button is a placeholder for that future "
            "integration - once available, it will compare local Floor Price/MSRP/Cost/"
            "Current Cost against Jeeves' own prices and generate a difference report, "
            "exportable and re-importable here to adjust prices."
        )

    if action == "acknowledge_warning":
        warning_id = request.form.get("warning_id", "")
        warnings = load_or_create_purchasing_warnings()
        for w in warnings:
            if w["id"] == warning_id:
                w["acknowledged"] = True
                w["acknowledged_at"] = datetime.now().isoformat(timespec="seconds")
        save_purchasing_warnings(warnings)

    open_warnings = [w for w in load_or_create_purchasing_warnings() if not w["acknowledged"]]

    flagged = [
        p for p in parts
        if any(p.get(f) in (None, "") for f in PRICE_FIELDS)
    ]

    report_generated = None
    if action == "generate_catalog_report":
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        report_path = REPORTS_DIR / f"purchasing_catalog_report_{datetime.now():%Y%m%d_%H%M%S}.csv"
        fieldnames = ["brand", "platform", "category", "code", "description"] + PRICE_FIELDS
        with open(report_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for p in flagged:
                writer.writerow({k: p.get(k) for k in fieldnames})
        report_generated = report_path.name

    action_items = compute_quote_action_items(parts, quotes)

    quotes_report_generated = None
    if action == "generate_quotes_report":
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        report_path = REPORTS_DIR / f"purchasing_quotes_report_{datetime.now():%Y%m%d_%H%M%S}.csv"
        fieldnames = ["display_id", "opportunity_id", "locked", "brand", "platform", "category", "code", "description", "missing"]
        with open(report_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for item in action_items:
                writer.writerow({k: item[k] for k in fieldnames})
        quotes_report_generated = report_path.name

    dashboard_counts = compute_quote_action_item_counts(action_items)

    return render_template(
        "purchasing.html",
        flagged=flagged,
        total_parts=len(parts),
        report_generated=report_generated,
        pending_import=pending_import,
        jeeves_message=jeeves_message,
        price_fields=PRICE_FIELDS,
        action_items=action_items,
        quotes_report_generated=quotes_report_generated,
        total_quotes=len(quotes),
        upload_result=upload_result,
        dashboard_counts=dashboard_counts,
        open_warnings=open_warnings,
    )


@app.route("/purchasing/pricing_gaps")
def purchasing_pricing_gaps():
    """Read-only view of §1's flagged parts - added 2026-08-18 to replace the
    old inline-editable table (which was also the giant ~17K-field form that
    caused a real 413 earlier the same day). No form, no inputs - just a
    look. Editing happens via Export Pricing Sheet -> edit offline -> Upload."""
    parts = load_parts()
    flagged = [
        p for p in parts
        if any(p.get(f) in (None, "") for f in PRICE_FIELDS)
    ]
    return render_template(
        "purchasing_pricing_gaps.html",
        flagged=flagged,
        total_parts=len(parts),
        price_fields=PRICE_FIELDS,
    )


@app.route("/purchasing/download/<path:filename>")
def purchasing_download(filename):
    """Serves a previously-generated report out of data/reports/ as a real
    browser download - added 2026-08-18 because Generate Catalog Report/
    Generate Report used to only write the file to the server's local disk
    and show its filename as plain text, with no way to actually get it
    from the browser. Path is resolved and checked against REPORTS_DIR
    to block '../' traversal outside the reports folder."""
    safe_path = (REPORTS_DIR / filename).resolve()
    if REPORTS_DIR.resolve() not in safe_path.parents or not safe_path.is_file():
        return "Report not found.", 404
    return send_file(safe_path, as_attachment=True, download_name=filename)


@app.route("/admin", methods=["GET", "POST"])
def admin():
    parts = load_parts()
    quotes = load_quotes()
    approvals = load_approvals()

    action_items = compute_quote_action_items(parts, quotes)
    unreviewed = compute_unreviewed_base_models(parts, approvals)
    open_warnings_count = len([w for w in load_or_create_purchasing_warnings() if not w["acknowledged"]])

    SAMPLE_TEST_CUSTOMERS = [
        "Acme Manufacturing", "Northwind Logistics", "Sunrise Distribution",
        "Blue Ridge Industrial", "Harborview Freight",
    ]
    if request.method == "POST" and request.form.get("action") == "seed_test_customers":
        customers = load_customers()
        existing_names = {c["name"] for c in customers}
        now = datetime.now().isoformat(timespec="seconds")
        for name in SAMPLE_TEST_CUSTOMERS:
            if name not in existing_names:
                customers.append({"name": name, "source": "test", "hubspot_id": None, "created_at": now})
        save_customers(customers)

    if request.method == "POST" and request.form.get("action") == "remove_test_customers":
        customers = [c for c in load_customers() if c.get("source") != "test"]
        save_customers(customers)

    customers = load_customers()
    unlinked_customers = compute_unlinked_customers(customers)
    test_customers = sorted([c for c in customers if c.get("source") == "test"], key=lambda c: c["name"])

    customer_report_generated = None
    if request.method == "POST" and request.form.get("action") == "generate_customer_report":
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        report_path = REPORTS_DIR / f"unlinked_customers_report_{datetime.now():%Y%m%d_%H%M%S}.csv"
        fieldnames = ["name", "source", "hubspot_id", "created_at"]
        with open(report_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for c in unlinked_customers:
                writer.writerow({k: c.get(k) for k in fieldnames})
        customer_report_generated = report_path.name

    rep_error = None
    if request.method == "POST" and request.form.get("action") == "add_rep":
        rep_name = request.form.get("rep_name", "").strip()
        rep_code = request.form.get("rep_code", "").strip()
        if not rep_name:
            rep_error = "Rep name is required."
        elif not (rep_code.isdigit() and len(rep_code) == 4):
            rep_error = "Code must be exactly 4 digits."
        else:
            reps = load_sales_reps()
            if any(r["name"] == rep_name for r in reps):
                rep_error = f'"{rep_name}" is already in the list.'
            else:
                reps.append({
                    "name": rep_name, "code": rep_code, "locked": False,
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                })
                save_sales_reps(reps)

    if request.method == "POST" and request.form.get("action") == "remove_rep":
        rep_name = request.form.get("rep_name", "").strip()
        reps = [r for r in load_sales_reps() if r["name"] != rep_name]
        save_sales_reps(reps)

    # Toggles a rep between usable and locked - a locked rep can't be picked
    # on Sales (excluded from /api/sales_reps) or verify/save/copy a quote
    # (rep_is_usable() below), but their past quotes stay attributed to them
    # unchanged (created_by/sales_rep are frozen strings on the quote, not a
    # live reference to the rep record).
    if request.method == "POST" and request.form.get("action") == "toggle_rep_lock":
        rep_name = request.form.get("rep_name", "").strip()
        reps = load_sales_reps()
        for r in reps:
            if r["name"] == rep_name:
                r["locked"] = not r.get("locked", False)
        save_sales_reps(reps)

    reset_pin_result = None
    if request.method == "POST" and request.form.get("action") == "reset_rep_pin":
        rep_name = request.form.get("rep_name", "").strip()
        reps = load_sales_reps()
        new_code = "".join(secrets.choice("0123456789") for _ in range(4))
        for r in reps:
            if r["name"] == rep_name:
                r["code"] = new_code
                reset_pin_result = {"name": rep_name, "code": new_code}
        save_sales_reps(reps)

    pin_error = None
    if request.method == "POST" and request.form.get("action") == "change_site_pin":
        new_pin = request.form.get("new_pin", "").strip()
        if not (new_pin.isdigit() and 4 <= len(new_pin) <= 8):
            pin_error = "PIN must be 4-8 digits."
        else:
            access = load_or_create_site_access()
            access["pin"] = new_pin
            save_site_access(access)

    purchasing_pin_error = None
    if request.method == "POST" and request.form.get("action") == "change_purchasing_pin":
        new_pin = request.form.get("new_purchasing_pin", "").strip()
        if not (new_pin.isdigit() and len(new_pin) == 4):
            purchasing_pin_error = "PIN must be exactly 4 digits."
        else:
            access = load_or_create_site_access()
            access["purchasing_pin"] = new_pin
            save_site_access(access)

    all_platforms = sorted({p["platform"] for p in parts})

    return render_template(
        "admin.html",
        total_quotes=len(quotes),
        action_items_count=len(action_items),
        unreviewed=unreviewed,
        reviewed_count=len(all_platforms) - len(unreviewed),
        total_platforms=len(all_platforms),
        unlinked_customers=unlinked_customers,
        customer_report_generated=customer_report_generated,
        test_customers=test_customers,
        sales_reps=sorted(load_sales_reps(), key=lambda r: r["name"]),
        rep_error=rep_error,
        reset_pin_result=reset_pin_result,
        site_pin=load_or_create_site_access()["pin"],
        pin_error=pin_error,
        purchasing_pin=load_or_create_site_access()["purchasing_pin"],
        purchasing_pin_error=purchasing_pin_error,
        open_warnings_count=open_warnings_count,
    )


# ------------------------------------------------------------------ quote API

@app.route("/api/quotes")
def api_quotes_for_opportunity():
    opportunity_id = request.args.get("opportunity_id", "")
    quotes = load_quotes()
    matches = [q for q in quotes.values() if q["opportunity_id"] == opportunity_id]
    matches.sort(key=lambda q: (q["quote_number"], q["rev_number"]))
    return jsonify([
        {
            "display_id": display_id(q),
            "opportunity_id": q["opportunity_id"],
            "customer": q.get("customer", ""),
            "quote_number": q["quote_number"],
            "rev_number": q["rev_number"],
            "locked": q["locked"],
            "brand": q.get("brand", "JLT"),
            "platform": q["platform"],
            "part_number": q["part_number"],
            "updated_at": q["updated_at"],
        }
        for q in matches
    ])


@app.route("/api/quotes/all")
def api_quotes_all():
    """Every saved quote, for the Sales-page 'Lookup saved quote' panel -
    unlike /api/quotes this isn't filtered to one Opportunity ID.

    Optional ?customer=<name>: scopes results the way a real HubSpot search
    would once that connector exists, per the user 2026-08-17. Only sent by
    the client when the currently active customer was found via Customer
    Lookup (simulating a real HubSpot pull - see sales.html's
    `customerSource`), not for a manually-typed one. When present, returns
    only (a) that customer's own quotes, and (b) quotes belonging to
    customers not yet linked to HubSpot (`source: "manual"` in
    customers.json - the same customers Admin's "pending HubSpot link"
    report tracks) - a real search wouldn't surface some OTHER unrelated
    customer's quotes, but SHOULD surface orphaned ones that might actually
    belong to the customer just found, pending review. Quotes belonging to
    a *different* named customer are excluded entirely."""
    customer_filter = (request.args.get("customer") or "").strip()
    quotes = load_quotes()
    out = [
        {
            "display_id": display_id(q),
            "opportunity_id": q["opportunity_id"],
            "customer": q.get("customer", ""),
            "quote_number": q["quote_number"],
            "rev_number": q["rev_number"],
            "locked": q["locked"],
            "brand": q.get("brand", "JLT"),
            "platform": q["platform"],
            "part_number": q["part_number"],
            "updated_at": q["updated_at"],
        }
        for q in quotes.values()
    ]
    if customer_filter:
        manual_names = {c["name"] for c in load_customers() if c.get("source") == "manual"}
        out = [q for q in out if q["customer"] == customer_filter or q["customer"] in manual_names]
    out.sort(key=lambda q: q["updated_at"], reverse=True)
    return jsonify(out)


@app.route("/api/customers")
def api_customers_list():
    """Powers the Sales-page Customer Lookup panel, which simulates
    searching HubSpot (see sales.html's customerSource handling - there's no
    real connector yet). A customer created via Manual Customer represents
    someone confirmed NOT to be in HubSpot, so a real HubSpot search
    wouldn't find them either - excluded here so Lookup can't hand back a
    "manual" record and simulate finding it in HubSpot. source:"test" stays
    included on purpose: those are seeded specifically to exercise this
    panel (see Admin's "Add 5 Test Customers"), standing in for what a real
    HubSpot result would look like. This does NOT affect Admin's "pending
    HubSpot link" report, which reads customers.json directly server-side,
    not through this endpoint."""
    query = (request.args.get("q") or "").strip().lower()
    customers = [c for c in load_customers() if c.get("source") != "manual"]
    if query:
        customers = [c for c in customers if query in c["name"].lower()]
    out = [{"name": c["name"], "source": c.get("source", "manual")} for c in customers]
    out.sort(key=lambda c: c["name"])
    return jsonify(out)


@app.route("/api/customers", methods=["POST"])
def api_customers_create():
    """Manual Customer button: records a customer with source='manual' and no
    hubspot_id. These show up on the Admin page as pending a HubSpot link, so
    a manually-typed name doesn't silently become an orphaned opportunity
    once the real connector exists."""
    body = request.get_json(force=True)
    name = (body.get("name") or "").strip()
    if not name:
        return jsonify({"error": "Customer name is required."}), 400
    customers = load_customers()
    if not any(c["name"] == name for c in customers):
        customers.append({
            "name": name,
            "source": "manual",
            "hubspot_id": None,
            "created_at": datetime.now().isoformat(timespec="seconds"),
        })
        save_customers(customers)
    return jsonify({"name": name})


@app.route("/api/sales_reps")
def api_sales_reps_list():
    # Locked reps are excluded from the picker entirely - see rep_code_matches().
    names = sorted(r["name"] for r in load_sales_reps() if not r.get("locked"))
    return jsonify(names)


@app.route("/api/sales_reps/verify", methods=["POST"])
def api_sales_reps_verify():
    body = request.get_json(force=True)
    name = (body.get("name") or "").strip()
    code = (body.get("code") or "").strip()
    reps = load_sales_reps()
    match = next((r for r in reps if r["name"] == name), None)
    if match is None:
        return jsonify({"ok": False, "error": "Unknown rep."}), 404
    if match.get("locked"):
        return jsonify({"ok": False, "error": "This rep is locked. Contact an admin."}), 403
    if match["code"] != code:
        return jsonify({"ok": False, "error": "Code doesn't match."}), 401
    return jsonify({"ok": True, "name": name})


@app.route("/api/quotes/<path:opportunity_id>/<int:quote_number>")
def api_quote_get(opportunity_id, quote_number):
    quotes = load_quotes()
    q = quotes.get(lineage_key(opportunity_id, quote_number))
    if q is None:
        return jsonify({"error": "not found"}), 404
    out = dict(q)
    out["display_id"] = display_id(q)
    return jsonify(out)


@app.route("/api/quotes/<path:opportunity_id>/<int:quote_number>/revisions")
def api_quote_revisions(opportunity_id, quote_number):
    """Backs the revision browser next to Copy to New Opportunity (added
    2026-08-25). Returns the real stored history for quotes saved since
    revision_snapshot() started running - a quote saved entirely before
    that will only show whatever revisions accumulated from the next time
    it was edited onward, since nothing earlier was ever kept to recover."""
    quotes = load_quotes()
    q = quotes.get(lineage_key(opportunity_id, quote_number))
    if q is None:
        return jsonify({"error": "not found"}), 404
    revisions = q.get("revisions") or [revision_snapshot(q)]
    out = []
    for r in revisions:
        r = dict(r)
        r["display_id"] = f"{opportunity_id}-{quote_number}-{r['rev_number']}"
        out.append(r)
    return jsonify(out)


@app.route("/api/quotes/save", methods=["POST"])
def api_quote_save():
    body = request.get_json(force=True)
    opportunity_id = (body.get("opportunity_id") or "").strip()
    if not opportunity_id:
        return jsonify({"error": "Opportunity ID is required."}), 400

    sales_rep = (body.get("sales_rep") or "").strip()
    sales_rep_code = (body.get("sales_rep_code") or "").strip()
    if not sales_rep:
        return jsonify({"error": "Select a Sales Rep first."}), 400
    rep_match = next((r for r in load_sales_reps() if r["name"] == sales_rep), None)
    if not rep_code_matches(rep_match, sales_rep_code):
        return jsonify({"error": "Sales rep code doesn't match. Re-enter your 4-digit code."}), 401

    customer = (body.get("customer") or "").strip()
    brand = body.get("brand") or "JLT"
    platform = body.get("platform")
    selections = body.get("selections") or {}
    quote_number = body.get("quote_number")

    parts = load_parts()
    lines = build_snapshot(parts, brand, platform, selections)
    floor_total, msrp_total = quote_totals(lines)
    part_number = quote_part_number(platform, lines)
    now = datetime.now().isoformat(timespec="seconds")

    quotes = load_quotes()

    if quote_number is not None:
        key = lineage_key(opportunity_id, quote_number)
        q = quotes.get(key)
        if q is None:
            return jsonify({"error": "Quote not found."}), 404
        if q["locked"]:
            return jsonify({"error": "This quote is locked. Unlock it before making changes."}), 400
        # Backfill for a quote saved before revision history existed
        # (2026-08-25) - captures the state about to be superseded so
        # everything from this point forward is real history, even though
        # whatever was already overwritten before this feature existed
        # can't be recovered. A no-op for any quote that already has this.
        if "revisions" not in q:
            q["revisions"] = [revision_snapshot(q)]
        # Rev bumps on any REAL content change to an already-saved quote,
        # regardless of lock history - the original rule ("only bump if the
        # quote was ever locked") didn't match what a rep actually expects:
        # reported live 2026-08-17, load a quote, change an option, Accept,
        # Save - and the rev stayed put because it had never been locked.
        # Comparing against what's already stored (not just "was Save
        # clicked") avoids bumping on a no-op re-save (e.g. Accept+Save
        # clicked twice with nothing actually changed in between).
        config_changed = (
            lines != q.get("selections")
            or brand != q.get("brand")
            or platform != q.get("platform")
        )
        if config_changed:
            q["rev_number"] += 1
        q["customer"] = customer
        q["brand"] = brand
        q["platform"] = platform
        q["selections"] = lines
        q["floor_total"] = floor_total
        q["msrp_total"] = msrp_total
        q["part_number"] = part_number
        q["sales_rep"] = sales_rep
        q["updated_at"] = now
        if config_changed:
            q["revisions"].append(revision_snapshot(q))
    else:
        quote_number = next_quote_number(quotes, opportunity_id)
        key = lineage_key(opportunity_id, quote_number)
        q = {
            "opportunity_id": opportunity_id,
            "customer": customer,
            "quote_number": quote_number,
            "rev_number": 0,
            "locked": False,
            "ever_locked": False,
            "brand": brand,
            "platform": platform,
            "selections": lines,
            "floor_total": floor_total,
            "msrp_total": msrp_total,
            "part_number": part_number,
            "created_by": sales_rep,
            "sales_rep": sales_rep,
            "created_at": now,
            "updated_at": now,
        }
        q["revisions"] = [revision_snapshot(q)]
        quotes[key] = q

    # Every save locks the quote (added 2026-08-17, per the user) - a saved
    # quote is treated as "this is what I'm proposing right now", protected
    # from accidental further edits. To make another change, Unlock first
    # (the button is enabled again immediately - dirty is false right after
    # a fresh save), edit, Accept, Save - which locks it again. Applies to
    # the very first save too, not just revisions.
    q["locked"] = True
    q["ever_locked"] = True

    save_quotes(quotes)
    out = dict(q)
    out["display_id"] = display_id(q)
    return jsonify(out)


@app.route("/api/quotes/<path:opportunity_id>/<int:quote_number>/lock", methods=["POST"])
def api_quote_lock(opportunity_id, quote_number):
    body = request.get_json(force=True)
    locked = bool(body.get("locked"))

    quotes = load_quotes()
    key = lineage_key(opportunity_id, quote_number)
    q = quotes.get(key)
    if q is None:
        return jsonify({"error": "Quote not found."}), 404

    q["locked"] = locked
    if locked:
        q["ever_locked"] = True
    q["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_quotes(quotes)

    out = dict(q)
    out["display_id"] = display_id(q)
    return jsonify(out)


@app.route("/api/quotes/<path:opportunity_id>/<int:quote_number>/copy", methods=["POST"])
def api_quote_copy(opportunity_id, quote_number):
    body = request.get_json(force=True)
    new_opportunity_id = (body.get("new_opportunity_id") or "").strip()
    new_customer = (body.get("new_customer") or "").strip()
    if not new_opportunity_id:
        return jsonify({"error": "New Opportunity ID is required."}), 400

    sales_rep = (body.get("sales_rep") or "").strip()
    sales_rep_code = (body.get("sales_rep_code") or "").strip()
    if not sales_rep:
        return jsonify({"error": "Select a Sales Rep first."}), 400
    rep_match = next((r for r in load_sales_reps() if r["name"] == sales_rep), None)
    if not rep_code_matches(rep_match, sales_rep_code):
        return jsonify({"error": "Sales rep code doesn't match. Re-enter your 4-digit code."}), 401

    quotes = load_quotes()
    src = quotes.get(lineage_key(opportunity_id, quote_number))
    if src is None:
        return jsonify({"error": "Source quote not found."}), 404

    new_quote_number = next_quote_number(quotes, new_opportunity_id)
    now = datetime.now().isoformat(timespec="seconds")
    new_q = {
        "opportunity_id": new_opportunity_id,
        "customer": new_customer or src.get("customer", ""),
        "quote_number": new_quote_number,
        "rev_number": 0,
        "locked": False,
        "ever_locked": False,
        "brand": src.get("brand", "JLT"),
        "platform": src["platform"],
        "selections": [dict(l) for l in src["selections"]],
        "floor_total": src["floor_total"],
        "msrp_total": src["msrp_total"],
        "part_number": src["part_number"],
        "created_by": sales_rep,
        "sales_rep": sales_rep,
        "copied_from": display_id(src),
        "created_at": now,
        "updated_at": now,
    }
    new_q["revisions"] = [revision_snapshot(new_q)]
    quotes[lineage_key(new_opportunity_id, new_quote_number)] = new_q
    save_quotes(quotes)

    out = dict(new_q)
    out["display_id"] = display_id(new_q)
    return jsonify(out)


@app.route("/api/quotes/<path:opportunity_id>/<int:quote_number>/upload", methods=["POST"])
def api_quote_upload(opportunity_id, quote_number):
    return jsonify({
        "status": "not_connected",
        "message": "HubSpot isn't connected yet. This button is a placeholder for that future integration.",
    })


def build_quote_workbook(q):
    """Builds the same .xlsx export both the Sales-page download and the
    (currently dormant, see HANDOFF.md §9) HubSpot pricing-export attach
    route use - pulled out on 2026-08-25 so both share one implementation
    instead of the HubSpot path duplicating this."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Quote"

    bold = Font(bold=True)
    ws["A1"] = "Quote ID:"
    ws["B1"] = display_id(q)
    ws["A2"] = "Sales Rep:"
    ws["B2"] = q.get("sales_rep", "")
    ws["A3"] = "Customer:"
    ws["B3"] = q.get("customer", "")
    ws["A4"] = "Opportunity ID:"
    ws["B4"] = q["opportunity_id"]
    ws["A5"] = "Brand:"
    ws["B5"] = q.get("brand", "JLT")
    ws["A6"] = "Platform:"
    ws["B6"] = q["platform"]
    ws["A7"] = "Part Number:"
    ws["B7"] = q["part_number"]
    ws["A8"] = "Date:"
    ws["B8"] = q["updated_at"]
    for cell in ("A1", "A2", "A3", "A4", "A5", "A6", "A7", "A8"):
        ws[cell].font = bold

    header_row = 10
    headers = ["Category", "Code", "Description"] + CUSTOMER_FACING_PRICE_FIELDS
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = bold

    r = header_row + 1
    for line in q["selections"]:
        ws.cell(row=r, column=1, value=line["category"])
        ws.cell(row=r, column=2, value=line["code"])
        ws.cell(row=r, column=3, value=line["description"])
        ws.cell(row=r, column=4, value=line["Floor Price"])
        ws.cell(row=r, column=5, value=line["MSRP"])
        r += 1

    ws.cell(row=r + 1, column=3, value="Floor Total").font = bold
    ws.cell(row=r + 1, column=4, value=q["floor_total"])
    ws.cell(row=r + 2, column=3, value="MSRP Total").font = bold
    ws.cell(row=r + 2, column=4, value=q["msrp_total"])

    for col, width in zip("ABCDE", (24, 10, 50, 14, 14)):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@app.route("/api/quotes/<path:opportunity_id>/<int:quote_number>/export.xlsx")
def api_quote_export(opportunity_id, quote_number):
    quotes = load_quotes()
    q = quotes.get(lineage_key(opportunity_id, quote_number))
    if q is None:
        return "Quote not found", 404

    buf = build_quote_workbook(q)
    filename = f"quote_{display_id(q)}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# --------------------------------------------------------- HubSpot (dormant)
# Added 2026-08-25, per HANDOFF.md §9 - NOT wired to any button or template.
# The Sales-page "Upload Hspt" button and the free-typed Opportunity ID field
# still point at the stub api_quote_upload() above and unchanged local
# customer/quote lookups, exactly as before. These routes exist so the code
# is ready and reviewable, but nothing calls them yet: there is no HubSpot
# Private App token configured (data/hubspot_config.json's access_token is
# still null), so every one of them would currently raise
# hubspot_client.HubSpotNotConfigured. Do not repoint any UI at these until
# a real token exists and each route has been tested against a live
# HubSpot account - see CHANGELOG.md's 2026-08-25 entry for exactly what's
# unverified (the association type IDs, and which quote total becomes the
# deal amount).

@app.route("/api/hubspot/customers")
def api_hubspot_customer_search():
    query = (request.args.get("q") or "").strip()
    if not query:
        return jsonify([])
    try:
        return jsonify(hubspot_client.search_customers(query))
    except hubspot_client.HubSpotNotConfigured as exc:
        return jsonify({"error": str(exc)}), 503
    except hubspot_client.HubSpotError as exc:
        return jsonify({"error": str(exc)}), 502


@app.route("/api/hubspot/companies/<company_id>/deals")
def api_hubspot_company_deals(company_id):
    try:
        return jsonify(hubspot_client.get_open_deals_for_company(company_id))
    except hubspot_client.HubSpotNotConfigured as exc:
        return jsonify({"error": str(exc)}), 503
    except hubspot_client.HubSpotError as exc:
        return jsonify({"error": str(exc)}), 502


@app.route("/api/hubspot/deals_for_customer")
def api_hubspot_deals_for_customer():
    """Backs the Sales page's "Query Hbst" button (added 2026-08-25) -
    takes just a customer name, since the browser doesn't track a HubSpot
    company ID today, and resolves name -> company -> open deals in one
    round trip."""
    name = (request.args.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name is required."}), 400
    try:
        return jsonify(hubspot_client.find_open_deals_for_customer_name(name))
    except hubspot_client.HubSpotNotConfigured as exc:
        return jsonify({"error": str(exc)}), 503
    except hubspot_client.HubSpotError as exc:
        return jsonify({"error": str(exc)}), 502


@app.route("/api/quotes/<path:opportunity_id>/<int:quote_number>/hubspot/push", methods=["POST"])
def api_quote_hubspot_push(opportunity_id, quote_number):
    quotes = load_quotes()
    q = quotes.get(lineage_key(opportunity_id, quote_number))
    if q is None:
        return jsonify({"error": "Quote not found."}), 404

    body = request.get_json(force=True)
    deal_id = (body.get("deal_id") or "").strip()
    if not deal_id:
        return jsonify({"error": "deal_id is required."}), 400

    try:
        result = hubspot_client.push_quote_to_deal(deal_id, q)
    except hubspot_client.HubSpotNotConfigured as exc:
        return jsonify({"error": str(exc)}), 503
    except hubspot_client.HubSpotError as exc:
        return jsonify({"error": str(exc)}), 502

    q["hubspot_deal_id"] = deal_id
    q["hubspot_line_item_ids"] = result["line_item_ids"]
    q["updated_at"] = datetime.now().isoformat(timespec="seconds")
    save_quotes(quotes)
    return jsonify(result)


def _record_hubspot_note(q, quotes, result, filename):
    notes = q.setdefault("hubspot_notes", [])
    notes.append({
        "note_id": result["note_id"],
        "file_id": result["file_id"],
        "filename": filename,
        "sent_at": datetime.now().isoformat(timespec="seconds"),
    })
    save_quotes(quotes)


@app.route(
    "/api/quotes/<path:opportunity_id>/<int:quote_number>/hubspot/attach_export",
    methods=["POST"],
)
def api_quote_hubspot_attach_export(opportunity_id, quote_number):
    """Interaction 4 in HANDOFF.md §9: attaches the same .xlsx the Sales
    page already lets a rep download - no separate file upload needed, the
    server builds it the same way build_quote_workbook() always has."""
    quotes = load_quotes()
    q = quotes.get(lineage_key(opportunity_id, quote_number))
    if q is None:
        return jsonify({"error": "Quote not found."}), 404

    body = request.get_json(force=True)
    deal_id = (body.get("deal_id") or "").strip()
    if not deal_id:
        return jsonify({"error": "deal_id is required."}), 400

    filename = f"quote_{display_id(q)}.xlsx"
    file_bytes = build_quote_workbook(q).read()
    try:
        result = hubspot_client.attach_file_to_deal(
            deal_id, file_bytes, filename,
            note_body=f"Pricing export attached from SalesConfig quote {display_id(q)}",
        )
    except hubspot_client.HubSpotNotConfigured as exc:
        return jsonify({"error": str(exc)}), 503
    except hubspot_client.HubSpotError as exc:
        return jsonify({"error": str(exc)}), 502

    _record_hubspot_note(q, quotes, result, filename)
    return jsonify(result)


@app.route(
    "/api/quotes/<path:opportunity_id>/<int:quote_number>/hubspot/attach_file",
    methods=["POST"],
)
def api_quote_hubspot_attach_file(opportunity_id, quote_number):
    """Interaction 5 in HANDOFF.md §9: attaches whatever file the rep
    supplies (their own finished customer-facing quote, any format) -
    this is what the "Upload Hspt" button would eventually call."""
    quotes = load_quotes()
    q = quotes.get(lineage_key(opportunity_id, quote_number))
    if q is None:
        return jsonify({"error": "Quote not found."}), 404

    deal_id = (request.form.get("deal_id") or "").strip()
    uploaded = request.files.get("file")
    if not deal_id or uploaded is None:
        return jsonify({"error": "deal_id and file are required."}), 400

    try:
        result = hubspot_client.attach_file_to_deal(
            deal_id, uploaded.read(), uploaded.filename,
            note_body=f"Final quote attached from SalesConfig quote {display_id(q)}",
        )
    except hubspot_client.HubSpotNotConfigured as exc:
        return jsonify({"error": str(exc)}), 503
    except hubspot_client.HubSpotError as exc:
        return jsonify({"error": str(exc)}), 502

    _record_hubspot_note(q, quotes, result, uploaded.filename)
    return jsonify(result)


if __name__ == "__main__":
    app.run(debug=True, port=5000, threaded=True)
