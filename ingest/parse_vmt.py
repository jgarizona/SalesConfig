"""
Ingest a JLT VMT price-book workbook into a normalized parts list (JSON).

Each platform tab (1014P, 1214N, 1214P, 6012, ..., Verso 15 2024) shares the
same layout: column A holds a category label (blank cells inherit the label
above), column B an option code, column C a description, and a price block
starting at the "Floor Price" header found near row 5.

Usage:
    python parse_vmt.py <path-to-xlsx> [--out parts.json]
"""

import argparse
import json
import sys
from pathlib import Path

import openpyxl

# Reference/summary tabs that are not platform configuration matrices.
NON_PLATFORM_SHEETS = {
    "Certified modems",
    "Service Contract",
    "Insights",
    "Product Key",
    "Accessories",
    "Ivanti",
}


def find_price_header(ws, search_rows=15, search_cols=20):
    """Locate the 'Floor Price' header cell; price columns follow immediately after."""
    for r in range(1, search_rows + 1):
        for c in range(1, search_cols + 1):
            if ws.cell(row=r, column=c).value == "Floor Price":
                return r, c
    return None


def parse_platform_sheet(ws, platform_name, brand):
    header = find_price_header(ws)
    if header is None:
        return []
    header_row, price_col = header

    price_labels = []
    c = price_col
    while True:
        label = ws.cell(row=header_row, column=c).value
        if not label:
            break
        price_labels.append(label)
        c += 1

    parts = []
    current_category = None
    for r in range(header_row + 1, ws.max_row + 1):
        category = ws.cell(row=r, column=1).value
        code = ws.cell(row=r, column=2).value
        description = ws.cell(row=r, column=3).value

        if category:
            current_category = str(category).strip()

        if code is None:
            # Sub-header / note row within a category (e.g. "Windows 11 Versions",
            # "DOCK MUST BE ORDERED") rather than a real selectable option.
            continue

        prices = {}
        for i, label in enumerate(price_labels):
            prices[label] = ws.cell(row=r, column=price_col + i).value

        parts.append({
            "brand": brand,
            "platform": platform_name,
            "category": current_category,
            "code": str(code).strip() if code is not None else None,
            "description": str(description).strip() if description is not None else None,
            **prices,
        })

    return parts


def parse_workbook(path, brand="JLT"):
    """brand defaults to JLT since that's the only vendor layout this parser
    understands today. Winmate/Getac/CyberLabs use different sheet layouts
    and will need their own parse_*.py once that work happens - passing a
    different brand here does NOT make this function understand their format,
    it only tags whatever it does parse (which will likely be garbage/empty
    for a non-JLT-shaped file)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    all_parts = []
    platform_sheets = [n for n in wb.sheetnames if n.strip() not in NON_PLATFORM_SHEETS]

    for name in platform_sheets:
        ws = wb[name]
        parts = parse_platform_sheet(ws, name.strip(), brand)
        all_parts.extend(parts)

    return all_parts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("--out", default=None, help="Output JSON path (default: alongside input, parts_<name>.json)")
    ap.add_argument("--brand", default="JLT", help="Vendor brand to tag every row with (default: JLT)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx_path)
    parts = parse_workbook(xlsx_path, brand=args.brand)

    out_path = Path(args.out) if args.out else xlsx_path.with_name(f"parts_{xlsx_path.stem}.json")
    out_path.write_text(json.dumps(parts, indent=2), encoding="utf-8")

    by_platform = {}
    for p in parts:
        by_platform.setdefault(p["platform"], 0)
        by_platform[p["platform"]] += 1

    print(f"Parsed {len(parts)} option rows across {len(by_platform)} platforms:")
    for platform, count in sorted(by_platform.items()):
        print(f"  {platform}: {count}")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    sys.exit(main())
