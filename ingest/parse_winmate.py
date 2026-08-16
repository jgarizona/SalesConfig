"""
Ingest a Winmate Master Price Book workbook into the same normalized parts
shape parse_vmt.py produces.

Same underlying pattern as JLT (one sheet per platform; column A = category,
blank inherits the label above; column B = code; column C = description;
a price block somewhere in the header row) - but NOT identical layout:

  - Header row position varies sheet to sheet (seen at rows 3, 4, and 5
    across real Winmate sheets), unlike JLT's fixed-ish position.
  - Price column order/labels vary: some sheets have Cost/Floor Price/MSRP
    consecutive, some have Cost then MSRP with no Floor Price column at all,
    and at least one sheet spells it "Floor" instead of "Floor Price".

So unlike parse_vmt.py's "find one anchor cell, read labels rightward until
blank", this scans the whole header-row candidate for each known price label
independently (label-driven, not position-anchored) and only uses whichever
labels it actually finds on that specific sheet.

Usage:
    python parse_winmate.py <path-to-xlsx> [--out parts.json]
"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

from category_map import to_canonical

# Reference/summary tabs, if any turn up - none observed in the real
# Winmate workbook, but kept for parity with parse_vmt.py and future-proofing.
NON_PLATFORM_SHEETS = set()

# canonical PRICE_FIELDS name -> normalized (lowercased, whitespace-collapsed)
# spellings seen (or plausible) in real Winmate sheets.
PRICE_LABEL_ALIASES = {
    "Floor Price": ["floor price", "floor"],
    "MSRP": ["msrp"],
    "Cost": ["cost"],
    "Current Cost": ["current cost", "curr cost", "currentcost"],
}


def _norm(v):
    return re.sub(r"\s+", " ", str(v).strip().lower()) if v is not None else ""


def find_price_columns(ws, search_rows=10, search_cols=24):
    """Scan the top of the sheet for known price-label header cells,
    independently of each other and of column order. Returns
    (header_row, {canonical_field: column}) for whichever row has the most
    matches, or (None, {}) if nothing recognizable was found."""
    alias_to_field = {}
    for field, aliases in PRICE_LABEL_ALIASES.items():
        for a in aliases:
            alias_to_field[a] = field

    best_row, best_cols = None, {}
    for r in range(1, search_rows + 1):
        found = {}
        for c in range(1, search_cols + 1):
            val = _norm(ws.cell(row=r, column=c).value)
            if val in alias_to_field:
                field = alias_to_field[val]
                found.setdefault(field, c)
        if len(found) > len(best_cols):
            best_row, best_cols = r, found

    if not best_cols:
        return None, {}
    return best_row, best_cols


def parse_platform_sheet(ws, platform_name, brand):
    header_row, price_cols = find_price_columns(ws)
    if header_row is None:
        return []

    all_aliases = {a for aliases in PRICE_LABEL_ALIASES.values() for a in aliases}
    price_columns = sorted(price_cols.values())
    first_price_col = min(price_columns) if price_columns else None

    parts = []
    current_category = None
    layout = "matrix"  # "matrix" = category/code/description in cols A/B/C.
    flat_category = None  # set once a second section (e.g. "Accessories") is found.

    # Some sheets reuse the same short code (X/A/0/1...) for more than one
    # independent choice under one inherited category label, with only a
    # blank row (no new label) separating them - e.g. MH4005's
    # "Data Collection:" silently contains a Barcode Reader sub-choice, a
    # Smart Card Reader sub-choice, a Fingerprint Reader sub-choice, and an
    # NFC Reader sub-choice, each numbered X/A or 0/1 independently. Without
    # this, the second sub-choice's rows would collide on (platform,
    # category, code) with the first and silently overwrite it on merge -
    # confirmed 2026-08-15 on real data. Suffix repeats " (2)", " (3)", ...
    # rather than losing them.
    variant_state = {}

    def resolve_category(base_category, code_str):
        state = variant_state.get(base_category)
        if state is None:
            variant_state[base_category] = (1, {code_str})
            return base_category
        variant_num, seen = state
        if code_str in seen:
            variant_num += 1
            seen = {code_str}
        else:
            seen = seen | {code_str}
        variant_state[base_category] = (variant_num, seen)
        return base_category if variant_num == 1 else f"{base_category} ({variant_num})"

    r = header_row + 1
    while r <= ws.max_row:
        # A second/later section (observed real example: "Accessories") repeats
        # the price-label header cells at the SAME columns further down the
        # sheet, but with a flatter code/description-only layout below it (no
        # category column) - detect that boundary rather than assume every
        # sheet has exactly one table.
        if any(_norm(ws.cell(row=r, column=c).value) in all_aliases for c in price_columns):
            label = None
            for c in range(1, first_price_col or 4):
                v = ws.cell(row=r, column=c).value
                if v:
                    label = str(v).strip()
                    break
            flat_category = to_canonical(label) if label else (current_category or platform_name)
            layout = "flat"
            r += 1
            continue

        if layout == "matrix":
            raw_category = ws.cell(row=r, column=1).value
            code = ws.cell(row=r, column=2).value
            description = ws.cell(row=r, column=3).value
            if raw_category:
                current_category = to_canonical(str(raw_category).strip())
            category_for_row = current_category
        else:
            code = ws.cell(row=r, column=1).value
            description = ws.cell(row=r, column=2).value
            category_for_row = flat_category

        if code is None:
            # Sub-header / note row within a section, not a real option.
            r += 1
            continue

        code_str = str(code).strip()
        category_for_row = resolve_category(category_for_row, code_str)

        prices = {}
        for field in ["Floor Price", "MSRP", "Cost", "Current Cost"]:
            col = price_cols.get(field)
            prices[field] = ws.cell(row=r, column=col).value if col else None

        parts.append({
            "brand": brand,
            "platform": platform_name,
            "category": category_for_row,
            "code": code_str,
            "description": str(description).strip() if description is not None else None,
            "requires_review": False,  # manufacturer's own official catalog
            **prices,
        })
        r += 1

    return parts


def parse_workbook(path, brand="Winmate"):
    wb = openpyxl.load_workbook(path, data_only=True)
    all_parts = []
    platform_sheets = [n for n in wb.sheetnames if n.strip() not in NON_PLATFORM_SHEETS]

    for name in platform_sheets:
        ws = wb[name]
        parts = parse_platform_sheet(ws, name.strip(), brand)
        all_parts.extend(parts)

    return all_parts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("--out", default=None, help="Output JSON path (default: alongside input, parts_<name>.json)")
    ap.add_argument("--brand", default="Winmate", help="Vendor brand to tag every row with (default: Winmate)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx_path)
    parts = parse_workbook(xlsx_path, brand=args.brand)

    out_path = Path(args.out) if args.out else xlsx_path.with_name(f"parts_{xlsx_path.stem}.json")
    out_path.write_text(json.dumps(parts, indent=2), encoding="utf-8")

    by_platform = {}
    skipped_sheets = []
    for p in parts:
        by_platform.setdefault(p["platform"], 0)
        by_platform[p["platform"]] += 1

    print(f"Parsed {len(parts)} option rows across {len(by_platform)} platforms:")
    for platform, count in sorted(by_platform.items()):
        print(f"  {platform}: {count}")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    sys.exit(main())
