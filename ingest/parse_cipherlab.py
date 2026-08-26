"""
Ingest a CipherLab price-book workbook into the normalized parts shape.

Replaces the earlier flat "One Page" parser (2026-08-25, per the user - the
old source file, `CipherLab Price Increase effective 4_10_2026 Product
List.xlsx`, was a price-*increase* list covering ~61 product families but
missing Base Unit rows for many of them, and was judged wrong/unreliable).
The new source (`CipherLab USA RS38 Price Book 8062025 with formula.xlsx`)
covers far fewer families - one sheet per device (RK26, RK95, RS36, RS38)
plus three flat license/service sheets (904R ReMoCloud, 90W WheeCare OS
upgrade licenses, 90R OCR license) - but is structurally much richer for the
device sheets: each one carries a real "legend" of per-position option codes
(Wireless/RAM+ROM/Barcode Reader/Camera/Battery/Package/GMS/Control Code,
plus Keypad on RK95), then a "Terminal Kit" table of the vendor's actual
released SKUs with each SKU's product code broken into those same per-
position codes column-for-column. Two reference-only tabs ("SKU", an
index of CipherLab's whole product line naming convention; "Warranty", a
service-plan price matrix for models not present as device sheets in this
file) are intentionally skipped - neither is this file's own device data.

Still no true build-your-own configurator: the Terminal Kit table lists only
the specific combinations CipherLab actually released (e.g. RS38 has 3 real
SKUs, not the dozens of combinations its legend positions could combine
into), so - consistent with the old parser and with parse_getac.py - each
row becomes one fixed "Base Unit:" record, not several independently-
selectable per-category options. The legend gives something the old flat
file never had though: a reliable per-SKU decomposition into structured
`attributes` (cpu/os/os_version/ram/storage/wireless, matching
app.py's ATTRIBUTE_CATEGORY_MAP so Search by Requirements can actually find
these platforms, plus a few CipherLab-only facets - scanner/camera/battery -
that aren't wired into search today but are harmless extra metadata if that
ever changes) instead of best-effort regexes over one flat description.
Below the Terminal Kit table, each sheet also lists real accessories
(cradles, batteries, adapters, holsters, etc.) under their own section
labels (col B only, e.g. "Cradle", "4BC-Healthcare") - these become
"Add On Options:" records under the same platform, same treatment the old
parser gave CipherLab's flat accessory rows.

Detected structurally, not by a fixed row/column map, since section
depth (and therefore which column each legend category lands in) differs
sheet to sheet - RK95 has an extra "Keypad Options" category RS38 doesn't,
shifting everything after it one column right:
  - A "legend" row with exactly one populated cell among columns C:N is a
    category label for whatever column its value sits in.
  - A row with exactly 3 populated cells among C:N, the middle one literally
    "-", is a `code -> description` legend entry for the column the code
    sits in (associated with whichever label row last claimed that column).
  - The header row where column B == "Product Code" ends the legend region.
  - After the header, a row with only column B populated is a section label
    (e.g. "Terminal Kit", "Cradle", "904R ReMoCloud"). Rows under a
    "Terminal Kit" section carry per-position codes in the legend's columns
    (decoded via the legend); every other section's rows are flat
    code/description/price accessory rows, description straight from
    column D. Price is always column O (-> MSRP) / column P (-> Floor
    Price), whichever the sheet calls them ("FloorPrice" vs "Discount
    Price" - same shape, just relabeled per sheet).
  - 904R (ReMoCloud), 90W (Android OS-upgrade licenses), and 90R (OCR
    license) are **not real systems** (per the user, 2026-08-25, after
    seeing them rendered as selectable "platforms" on Sales - genuinely
    wrong, not a simplification) - they're license/service SKUs that attach
    to a real device platform, the same as any other accessory. Each one's
    rows are redistributed onto the real device platforms (RK26/RK95/RS36/
    RS38, whichever device sheets this workbook actually has) as ordinary
    "Add On Options:" instead of becoming platforms of their own:
      - **904R and 90R apply to every real platform** (their own
        descriptions never name a specific model - 904R's cloud-management
        service and 90R's OCR activation key are device-agnostic) - each of
        their rows is duplicated once per real platform.
      - **90W is model-specific per row**, grouped under section-label rows
        naming which model each license is for (e.g. "RK95 android OS
        upgrade license"). A row is attached to whichever real platform's
        name appears in its nearest preceding section label; rows under a
        section naming a model this workbook has no device sheet for (RK25,
        RS35, RS51, Hera51 - an older product, per the user, not part of
        this configurator) are dropped entirely rather than invented as a
        new platform.
    None of these three sheets ever produces a "Base Unit:" row either way -
    they're licenses, not devices - which is also why the old brand-wide
    Search exclusion isn't needed for them: `api_search_options` in app.py
    already excludes any (brand, platform) with zero Base Unit rows from the
    dropdown pool, and since their rows now live under platforms that do
    have Base Unit rows, they're simply ordinary add-on options a rep might
    see once a real platform is picked - not a search dead end.

Usage:
    python parse_cipherlab.py <path-to-xlsx> [--out parts.json]
"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

from os_facets import extract_os_version, extract_os_edition
from storage_facets import extract_storage_capacity

EXCLUDED_SHEETS = {"SKU", "Warranty"}

# Not real systems - license/service SKUs redistributed onto real device
# platforms instead of becoming platforms of their own (see module
# docstring). BROADCAST_SHEETS' rows apply to every real platform;
# PREFIX_MATCHED_SHEETS' rows are matched to a real platform by name via
# their section label, dropped if no real platform matches.
BROADCAST_SHEETS = {"904R", "90R"}
PREFIX_MATCHED_SHEETS = {"90W"}

CODE_COL = 2       # B
DESC_COL = 4       # D
LEGEND_MIN_COL = 3  # C
LEGEND_MAX_COL = 14  # N
PRICE_COL = 15      # O -> MSRP / List Price
FLOOR_COL = 16       # P -> Floor Price / Discount Price

_CPU_RE = re.compile(r"Qualcomm\s+\S+(?:\s+Octa-Core\s*[\d.]+\s*GHz)?", re.IGNORECASE)
_DISPLAY_RE = re.compile(r'(\d+(?:\.\d+)?)"\s*([A-Za-z0-9+]*)')
_RAM_RE = re.compile(r"(\d+\s*GB)\s*RAM", re.IGNORECASE)
_STORAGE_RE = re.compile(r"(\d+\s*GB)\s*Flash(?:\s*ROM)?\b", re.IGNORECASE)


def clean_text(v):
    if v is None:
        return None
    s = str(v).replace("_x000D_", " ").replace("_x000A_", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return s or None


def attr_key_for_label(label):
    """Best-effort map of a legend category label to an attributes-dict key.
    Substring match (not exact), since wording varies slightly sheet to
    sheet ('Wireless Communication Options' vs '...with NFC / SIM Options').
    Returns None for legend categories with nothing worth surfacing as
    search metadata (package/power-cord options, the cosmetic control
    code)."""
    l = label.lower()
    if "wireless" in l:
        return "wireless"
    if "ram" in l or "rom" in l:
        return "ram"
    if "barcode" in l or "reader" in l:
        return "scanner"
    if "camera" in l:
        return "camera"
    if "battery" in l:
        return "battery"
    if "keypad" in l:
        return "keypad"
    if "gms" in l:
        return "os"
    return None


def parse_platform_base_attributes(description):
    """Pulls cpu/os/os_version/display out of a platform's own base
    description row (e.g. 'Android 13, Qualcomm 4490 Octa-Core 2.4GHz, 8GB
    RAM + 128GB Flash ROM, 6" FHD+ with Capacitive Touch Panel...') - applied
    to every SKU on that platform as a baseline, then overridden per-SKU by
    whatever the legend actually resolves for that row (see
    parse_terminal_kit_row)."""
    attrs = {}
    if not description:
        return attrs
    m = _CPU_RE.search(description)
    if m:
        attrs["cpu"] = m.group(0).strip().rstrip(",.")
    os_version = extract_os_version(description)
    if os_version:
        attrs["os"] = os_version
        attrs["os_version"] = os_version
    os_edition = extract_os_edition(description)
    if os_edition:
        attrs["os_edition"] = os_edition
    m = _DISPLAY_RE.search(description)
    if m:
        size, extra = m.group(1), (m.group(2) or "").strip()
        attrs["display"] = f'{size}" {extra}'.strip()
    m = _RAM_RE.search(description)
    if m:
        attrs["ram"] = m.group(1).replace(" ", "")
    m = _STORAGE_RE.search(description)
    if m:
        cap = extract_storage_capacity(m.group(0))
        if cap:
            attrs["storage"] = cap
    return attrs


def find_legend_regions(ws, header_row):
    """Scans rows 1..header_row-1 for category-label and code-legend rows
    (see module docstring). Returns {column: {"label": str, "codes": {code:
    description}}}."""
    legend = {}
    current_label_col = {}
    for r in range(1, header_row):
        populated = {}
        for c in range(LEGEND_MIN_COL, LEGEND_MAX_COL + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                populated[c] = v
        if len(populated) == 1:
            col, val = next(iter(populated.items()))
            label = clean_text(val)
            if label:
                legend.setdefault(col, {"label": label, "codes": {}})
                legend[col]["label"] = label
        elif len(populated) == 3:
            cols = sorted(populated)
            c0, c1, c2 = cols
            if str(populated[c1]).strip() == "-":
                code = clean_text(populated[c0])
                desc = clean_text(populated[c2])
                if code is not None and desc is not None:
                    legend.setdefault(c0, {"label": None, "codes": {}})
                    legend[c0]["codes"][code] = desc
    return legend


def parse_terminal_kit_row(ws, row, legend, base_attrs):
    attrs = dict(base_attrs)
    desc_candidates = []
    for col, entry in legend.items():
        code = ws.cell(row=row, column=col).value
        if code in (None, ""):
            continue
        code = str(code).strip()
        option_desc = entry["codes"].get(code)
        if not option_desc:
            continue
        desc_candidates.append(option_desc)
        key = attr_key_for_label(entry["label"] or "")
        if key == "os":
            # GMS legend text states Android version directly, e.g.
            # "GMS(Android 13)" / "Regular Android OS (Non-GMS)(Android 13)"
            # - more precise than the platform-wide base description.
            os_version = extract_os_version(option_desc)
            if os_version:
                attrs["os"] = os_version
                attrs["os_version"] = os_version
        elif key == "ram":
            m = _RAM_RE.search(option_desc)
            if m:
                attrs["ram"] = m.group(1).replace(" ", "")
            m = _STORAGE_RE.search(option_desc)
            if m:
                cap = extract_storage_capacity(m.group(0))
                if cap:
                    attrs["storage"] = cap
        elif key:
            attrs[key] = option_desc

    # Full description: the longest populated text cell between the code
    # columns and the price column - the sheet's own combined-spec text.
    full_desc = None
    for c in range(LEGEND_MIN_COL, PRICE_COL):
        v = ws.cell(row=row, column=c).value
        if v is None:
            continue
        s = clean_text(v)
        if s and (full_desc is None or len(s) > len(full_desc)):
            full_desc = s
    if full_desc is None:
        full_desc = "; ".join(desc_candidates) or None

    return attrs, full_desc


def parse_sheet(ws, platform, brand):
    parts = []

    header_row = None
    for r in range(1, ws.max_row + 1):
        if clean_text(ws.cell(row=r, column=CODE_COL).value) == "Product Code":
            header_row = r
            break
    if header_row is None:
        return parts

    legend = find_legend_regions(ws, header_row)

    # Platform-wide base spec (e.g. row naming the base model 'AS38' next to
    # its full description) - whichever pre-header row has both a short
    # token in column B and a long description in column D.
    base_attrs = {}
    for r in range(1, header_row):
        b_val = clean_text(ws.cell(row=r, column=CODE_COL).value)
        d_val = clean_text(ws.cell(row=r, column=DESC_COL).value)
        if b_val and d_val and len(b_val) <= 12 and len(d_val) > 30:
            base_attrs = parse_platform_base_attributes(d_val)
            break

    current_section = None
    for r in range(header_row + 1, ws.max_row + 1):
        b_val = ws.cell(row=r, column=CODE_COL).value
        if b_val in (None, ""):
            continue
        code = clean_text(b_val)
        d_val = ws.cell(row=r, column=DESC_COL).value
        price = ws.cell(row=r, column=PRICE_COL).value
        rest_populated = any(
            ws.cell(row=r, column=c).value not in (None, "")
            for c in range(LEGEND_MIN_COL, LEGEND_MAX_COL + 1)
        )
        if price is None and d_val in (None, "") and not rest_populated:
            current_section = code
            continue

        if current_section == "Terminal Kit" and legend:
            attrs, description = parse_terminal_kit_row(ws, r, legend, base_attrs)
            category = "Base Unit:"
        else:
            attrs = dict(base_attrs)
            description = clean_text(d_val)
            category = "Add On Options:"

        floor = ws.cell(row=r, column=FLOOR_COL).value
        parts.append({
            "brand": brand,
            "platform": platform,
            "category": category,
            "code": code,
            "description": description,
            "requires_review": False,  # manufacturer's own official catalog
            "Floor Price": floor,
            "MSRP": price,
            "Cost": None,
            "Current Cost": None,
            "attributes": attrs,
        })

    return parts


def parse_service_rows(ws):
    """Flat rows for a 904R/90W/90R-style sheet - no legend, no Terminal
    Kit, just a header row then section-label rows (col B only) followed by
    code/description/price rows. Returns each row plus the section label it
    fell under, for the caller to decide which real platform(s) it belongs
    to (see BROADCAST_SHEETS/PREFIX_MATCHED_SHEETS)."""
    header_row = None
    for r in range(1, ws.max_row + 1):
        if clean_text(ws.cell(row=r, column=CODE_COL).value) == "Product Code":
            header_row = r
            break
    if header_row is None:
        return []

    rows = []
    current_section = None
    for r in range(header_row + 1, ws.max_row + 1):
        b_val = ws.cell(row=r, column=CODE_COL).value
        if b_val in (None, ""):
            continue
        code = clean_text(b_val)
        d_val = ws.cell(row=r, column=DESC_COL).value
        price = ws.cell(row=r, column=PRICE_COL).value
        rest_populated = any(
            ws.cell(row=r, column=c).value not in (None, "")
            for c in range(LEGEND_MIN_COL, LEGEND_MAX_COL + 1)
        )
        if price is None and d_val in (None, "") and not rest_populated:
            current_section = code
            continue
        rows.append({
            "code": code,
            "description": clean_text(d_val),
            "MSRP": price,
            "Floor Price": ws.cell(row=r, column=FLOOR_COL).value,
            "section": current_section or "",
        })
    return rows


def _service_part(brand, platform, row):
    return {
        "brand": brand,
        "platform": platform,
        "category": "Add On Options:",
        "code": row["code"],
        "description": row["description"],
        "requires_review": False,  # manufacturer's own official catalog
        "Floor Price": row["Floor Price"],
        "MSRP": row["MSRP"],
        "Cost": None,
        "Current Cost": None,
        "attributes": {},
    }


def parse_workbook(path, brand="CipherLab"):
    wb = openpyxl.load_workbook(path, data_only=True)

    device_sheets = [
        s for s in wb.sheetnames
        if s not in EXCLUDED_SHEETS and s not in BROADCAST_SHEETS and s not in PREFIX_MATCHED_SHEETS
    ]

    parts = []
    for sheet_name in device_sheets:
        parts.extend(parse_sheet(wb[sheet_name], platform=sheet_name, brand=brand))

    # 904R/90R: device-agnostic license/service SKUs - one copy under every
    # real platform (see module docstring).
    for sheet_name in wb.sheetnames:
        if sheet_name not in BROADCAST_SHEETS:
            continue
        for row in parse_service_rows(wb[sheet_name]):
            for platform in device_sheets:
                parts.append(_service_part(brand, platform, row))

    # 90W: model-specific licenses - attach to whichever real platform's
    # name appears in the row's section label, drop if none does (a model
    # this workbook has no device sheet for).
    for sheet_name in wb.sheetnames:
        if sheet_name not in PREFIX_MATCHED_SHEETS:
            continue
        for row in parse_service_rows(wb[sheet_name]):
            section = row["section"].lower()
            target = next((p for p in device_sheets if p.lower() in section), None)
            if target is None:
                continue
            parts.append(_service_part(brand, target, row))

    return parts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("--out", default=None, help="Output JSON path (default: alongside input, parts_<name>.json)")
    ap.add_argument("--brand", default="CipherLab", help="Vendor brand to tag every row with (default: CipherLab)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx_path)
    parts = parse_workbook(xlsx_path, brand=args.brand)

    out_path = Path(args.out) if args.out else xlsx_path.with_name(f"parts_{xlsx_path.stem}.json")
    out_path.write_text(json.dumps(parts, indent=2), encoding="utf-8")

    by_platform = {}
    base_unit_count = 0
    attr_hits = {}
    for p in parts:
        by_platform.setdefault(p["platform"], 0)
        by_platform[p["platform"]] += 1
        if p["category"] == "Base Unit:":
            base_unit_count += 1
        for k, v in p["attributes"].items():
            if v:
                attr_hits[k] = attr_hits.get(k, 0) + 1

    print(f"Parsed {len(parts)} rows across {len(by_platform)} platforms ({base_unit_count} Base Unit rows):")
    for platform, count in sorted(by_platform.items()):
        print(f"  {platform}: {count}")
    print("Attribute hit rates (of all rows):")
    for k, n in sorted(attr_hits.items()):
        print(f"  {k}: {n}/{len(parts)}")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    sys.exit(main())
