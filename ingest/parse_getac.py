"""
Ingest a Getac Select MSRP workbook into the normalized parts shape.

Structurally nothing like JLT/Winmate: one flat sheet, 5 columns (Model, SKU
ID, Description, MSRP, Currency), 371 rows. Each row is a complete, already
fixed pre-built SKU - Getac sells specific configurations, not a pick-one-
option-per-category matrix the way JLT/Winmate do. Decomposing a row's
description into independently-selectable category options would imply
combinations Getac doesn't actually sell, so every row becomes a single
"Base Unit:" record (code = SKU ID) rather than several per-category rows.

Since there's no separate column per spec, this pulls a best-effort CPU,
OS, RAM, storage, display, and wireless value out of the free-text
description into an `attributes` dict purely so Search by Requirements has
something to match against (see app.py's ATTRIBUTE_CATEGORY_MAP) - these
are NOT selectable options, just search metadata on an otherwise fixed
SKU. 100% hit rate on all six across the real 370-row catalog as of
2026-08-17 (verified, not assumed) - if a future price-list refresh
introduces a genuinely new phrasing these regexes don't cover, a row will
just silently lack that one attribute rather than error.

Usage:
    python parse_getac.py <path-to-xlsx> [--out parts.json]
"""

import argparse
import json
import re
import sys
from pathlib import Path

import openpyxl

CPU_PATTERNS = [
    re.compile(r"(?:Intel|AMD)\s+.+?Processor", re.IGNORECASE),
    re.compile(r"Qualcomm\s+\S+", re.IGNORECASE),
]


def extract_cpu(description):
    if not description:
        return None
    for pat in CPU_PATTERNS:
        m = pat.search(description)
        if m:
            return m.group(0).strip()
    return None


def extract_os(description):
    if not description:
        return None
    m = re.search(r"Windows\s*11\s*(?:Pro|Professional)", description, re.IGNORECASE)
    if m:
        return "Windows 11 Pro"
    m = re.search(r"Android\s*\d+", description, re.IGNORECASE)
    if m:
        return m.group(0).strip()
    return None


def extract_ram(description):
    if not description:
        return None
    m = re.search(r"(\d+)\s*GB\s+RAM", description, re.IGNORECASE)
    return f"{m.group(1)}GB" if m else None


def extract_storage(description):
    if not description:
        return None
    m = re.search(r"(\d+)\s*(GB|TB)\s+(?:PCIe\s+SSD|Storage)", description, re.IGNORECASE)
    return f"{m.group(1)}{m.group(2).upper()}" if m else None


def extract_display(description):
    """Size + resolution + touch, e.g. '13.3" Full HD Touchscreen'. Size is
    pulled from whichever comma-separated clause mentions "webcam" - the
    source consistently pairs display size with the webcam mention, but the
    quote-mark character after the number is corrupted (mojibake) on
    several rows (e.g. S510AD), so this deliberately doesn't require a
    literal `"` - just the leading number in that clause."""
    if not description:
        return None
    size = None
    for segment in description.split(","):
        if "webcam" in segment.lower():
            m = re.search(r"(\d+\.?\d*)", segment)
            if m:
                size = m.group(1) + '"'
            break
    res_m = re.search(r"\b(Full HD|WUXGA|HD)\b", description, re.IGNORECASE)
    resolution = res_m.group(1) if res_m else None
    touch = bool(re.search(r"Touchscreen", description, re.IGNORECASE))
    parts = [p for p in (size, resolution, "Touchscreen" if touch else None) if p]
    return " ".join(parts) if parts else None


def extract_wireless(description):
    if not description:
        return None
    parts = []
    if re.search(r"\bWi[- ]?Fi\b", description, re.IGNORECASE):
        parts.append("WiFi")
    if re.search(r"\bBT\b", description):
        parts.append("BT")
    cell_m = re.search(r"\b(4G LTE|5G Sub-6|5G)\b", description, re.IGNORECASE)
    if cell_m:
        parts.append(cell_m.group(1))
    return " + ".join(parts) if parts else None


def parse_workbook(path, brand="Getac"):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]  # single-sheet workbook; sheet name has a stray encoding artifact

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip() if c.value else None for c in header_cells]

    parts = []
    for values in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        model = row.get("Model")
        sku = row.get("SKU ID")
        description = row.get("Description")
        msrp = row.get("MSRP")
        if not (model and sku):
            continue

        description = str(description).strip() if description is not None else None
        if description is not None:
            # Collapses stray whitespace runs, including non-breaking spaces
            # (U+00A0) that show up on some rows as an artifact of the
            # source spreadsheet being pasted from elsewhere - left alone,
            # "Intel Core\xa0i5-1335U Processor" (nbsp) and "Intel Core
            # i5-1335U Processor" (regular space) extract as two distinct
            # attribute values for the same real CPU, splitting one search
            # match into two and silently hiding whichever rows use the
            # nbsp form from a search on the regular-space form.
            description = re.sub(r"\s+", " ", description).strip()
        attributes = {}
        cpu = extract_cpu(description)
        if cpu:
            attributes["cpu"] = cpu
        os_name = extract_os(description)
        if os_name:
            attributes["os"] = os_name
        ram = extract_ram(description)
        if ram:
            attributes["ram"] = ram
        storage = extract_storage(description)
        if storage:
            attributes["storage"] = storage
        display = extract_display(description)
        if display:
            attributes["display"] = display
        wireless = extract_wireless(description)
        if wireless:
            attributes["wireless"] = wireless

        parts.append({
            "brand": brand,
            "platform": str(model).strip(),
            "category": "Base Unit:",
            "code": str(sku).strip(),
            "description": description,
            "requires_review": False,  # manufacturer's own official catalog
            "Floor Price": None,
            "MSRP": msrp,
            "Cost": None,
            "Current Cost": None,
            "attributes": attributes,
        })

    return parts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx_path")
    ap.add_argument("--out", default=None, help="Output JSON path (default: alongside input, parts_<name>.json)")
    ap.add_argument("--brand", default="Getac", help="Vendor brand to tag every row with (default: Getac)")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx_path)
    parts = parse_workbook(xlsx_path, brand=args.brand)

    out_path = Path(args.out) if args.out else xlsx_path.with_name(f"parts_{xlsx_path.stem}.json")
    out_path.write_text(json.dumps(parts, indent=2), encoding="utf-8")

    by_platform = {}
    attr_hits = {"cpu": 0, "os": 0, "ram": 0, "storage": 0, "display": 0, "wireless": 0}
    for p in parts:
        by_platform.setdefault(p["platform"], 0)
        by_platform[p["platform"]] += 1
        for k in attr_hits:
            if p["attributes"].get(k):
                attr_hits[k] += 1

    print(f"Parsed {len(parts)} SKUs across {len(by_platform)} platforms:")
    for platform, count in sorted(by_platform.items()):
        print(f"  {platform}: {count}")
    print("Attribute hit rates:")
    for k, n in attr_hits.items():
        print(f"  {k}: {n}/{len(parts)}")
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    sys.exit(main())
